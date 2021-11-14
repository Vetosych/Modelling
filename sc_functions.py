## Все вспомогательные функции для расчета скоркарт

import os
import scipy.stats as ss
import numpy as np
import pandas as pd
import datetime as dtime
import statsmodels.api as sm
import gc
from openpyxl import load_workbook
from woeBinningPandas import woe_binning
from sklearn.model_selection import RandomizedSearchCV
from openpyxl.styles import Alignment
from openpyxl.chart import BarChart, Reference, LineChart, ScatterChart, Series
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
#from oracle_load import load_data_from_oracle_2

## Объекты для форматирования

thin_border_top_bot = Border(top=Side(style='thin'), 
                             bottom=Side(style='thin'))

thin_border_top_bot_r = Border(top=Side(style='thin'), 
                             bottom=Side(style='thin'),
                              right=Side(style='thin'))

thin_border_top_bot_l = Border(top=Side(style='thin'), 
                             bottom=Side(style='thin'),
                              left=Side(style='thin'))

thin_border_top_bot_l_r = Border(top=Side(style='thin'), 
                                 bottom=Side(style='thin'),
                                 left=Side(style='thin'),
                                 right=Side(style='thin'))

thin_border_top = Border(top=Side(style='thin'))

thin_border_bot = Border(bottom=Side(style='thin'))

thin_border_bot_r = Border(bottom=Side(style='thin'),
                           right=Side(style='thin'))

thin_border_top_r = Border(top=Side(style='thin'),
                           right=Side(style='thin'))

thin_border_l_r = Border(left=Side(style='thin'),
                           right=Side(style='thin'))

thin_border_bot_l_r = Border(left=Side(style='thin'),
                             bottom=Side(style='thin'),
                             right=Side(style='thin'))

thin_border_r = Border(right=Side(style='thin'))

## download_csv_file
# Загрузка файла csv

def download_csv_file(filepath, mysep = ';', myindex_col = None):
    mylist = []    
    for chunk in pd.read_csv(filepath_or_buffer = filepath, 
                             sep = mysep, 
                             encoding ='cp1251',
                             engine = 'python', 
                             decimal=",", 
                             index_col = myindex_col,
                             error_bad_lines=False, 
                             chunksize = 10000):
        mylist.append(chunk)       
    df = pd.concat(mylist, axis = 0)   
    del mylist
    return df

## download_excel_file
# Загрузка файла excel

def download_excel_file(filepath,sheetname, \
                        converters = None):      
    df = pd.read_excel(io = filepath,
                       sheet_name = sheetname,
                       converters = converters)   
    return df

## get_attr_type
# Проверка типа переменных на интервальную/категориальную

def get_attr_type(attr_type):
    if attr_type in ['int32','int64','float32','float64']:
        return 'Интервальная'
    elif attr_type in ['object','category']:
        return 'Категориальная'
    else:
        return 'Другое'

## get_detail_check
# Проверка для воронки отсечений переменной для лонг листа
# Если заполнена детализация, то значит переменная не прошла

def get_detail_check(detail, help_flg = ''):    
    if str(help_flg) == '': # нет доп флага 
        if detail == '':
            return 1
        else:
            return 0
    else: # есть доп флаг
        if detail == '' and help_flg == 1: 
            return 1
        else:
            return 0
        
## get_IV_detail

def get_IV_detail(iv_value,iv_detail,iv_side_check,iv_cutoff):
    if iv_detail != '':
        return iv_detail
    elif iv_side_check == 'left' and iv_value < iv_cutoff:
        return '1.1. IV меньше, чем ' + str(iv_cutoff)
    elif iv_side_check == 'right' and iv_value > iv_cutoff:
        return '1.2. IV больше, чем ' + str(iv_cutoff)
    else:
        return iv_detail
    
## get_Trend_detail
        
def get_Trend_detail(has_trend,
                     PSI,PSI_cutoff,PSI_bin_cutoff,
                     VOL,VOL_cutoff,VOL_bin_cutoff,
                     bin_value,bin_v_cutoff,
                     no_category,
                     no_events):
    if has_trend == 0:
        return '2.1. Тренд: нарушение тренда у переменной на тестовой выборке'
    elif PSI == 0:
        PSI_TYPE = ''
        if PSI_bin_cutoff == False:
            PSI_TYPE = 'переменной'
        else:
            PSI_TYPE = 'бина в переменной'
        return '2.2. Тренд: PSI ' + PSI_TYPE + ' > ' + str(PSI_cutoff)
    elif VOL == 0:
        VOL_TYPE = ''
        if VOL_bin_cutoff == False:
            VOL_TYPE = 'переменной'
        else:
            VOL_TYPE = 'бина в переменной'
        return '2.3. Тренд: волатильность ' + VOL_TYPE + ' > ' + str(VOL_cutoff)
    elif no_events == 1:
        return '2.4. Тренд: на тестовой выборке в одном из бинов отсутствуют значения'
    elif no_category == 1:
        return '2.5. Тренд: в тестовой выборке присутствует категориальное значение, которого не было в тренировочной'
    elif bin_value == 1:
        return '2.6. Тренд: один из бинов переменной незначим: менее, чем ' + str(bin_v_cutoff) + '% от выборки'    
    else:
        return ''

## get_Trend_Unique_detail
    
def get_Trend_Unique_detail(trend_detail, attr_max_iv_check):
    if trend_detail == '' and attr_max_iv_check != 1:
        return '2.7. Тренд: для этой переменной уже существует более оптимальное бинирование' 
    else:
        return trend_detail
    
## get_PValue_detail
    
def get_PValue_detail(attr_name, this_attr_name, pvalue_detail, pvalue, pvalue_cutoff, iter_num):
    if attr_name == this_attr_name and pvalue > pvalue_cutoff and pvalue_detail == '':
        return '3.' + str(iter_num) + '. PValue больше, чем ' + str(pvalue_cutoff) + '; итерация ' + str(iter_num)
    else:
        return pvalue_detail
    
## get_Correl_detail
        
def get_Correl_detail(attr_name, this_attr_name, correl_detail, correl_attr_name, correl_type, correl_cutoff):
    if attr_name == this_attr_name and correl_detail == '':
        if correl_type == 'int':
            return '4.1. Корреляция интервальная c ' + correl_attr_name + ' больше, чем ' + str(correl_cutoff)
        elif correl_type == 'cat':
            return '4.2. Корреляция категориальная c ' + correl_attr_name + ' больше, чем ' + str(correl_cutoff)
        else:
            return correl_detail
    else:
        return correl_detail
    
## get_detail_from_excel
        
def get_detail_from_excel(detail):
    if str(detail).lower() == 'nan':
        return ''
    else:
        return str(detail)
       
## get_true_cat_attr_name
    
def get_true_cat_attr_name(attr_name):
    if str(attr_name) != '' and str(attr_name).lower() != 'nan':
        return str(attr_name) + ', '
    else:
        return ''
    
## get_true_woe_value
    
def get_true_woe_value(woe):
    if woe == 0 or woe is None:
        return 0
    else:
        return round(woe/100,6)
    
## get_true_int_cut
    
def get_true_int_cut(interval,zpt,scb,side):
    if side == 'left':
        return interval[1:zpt]
    elif side == 'right':
        return interval[zpt+2:scb]
    else:
        return ''
 
## long_list_final_check
    
def long_list_final_check(IV,trend,pvalue,corr,balance):
    if IV == 1 and trend == 1 and pvalue == 1 and corr == 1 and balance == 1:
        return 1
    else:
        return 0

## long_list_final_detail
    
def long_list_final_detail(IV,trend,pvalue,corr,balance):
    if IV != '':
        return IV
    elif trend != '':
        return trend
    elif pvalue != '':
        return pvalue
    elif corr != '':
        return corr
    elif balance != '':
        return balance
    else:
        return '99. Переменная участвует в модели'
    
## binning_by_param_loop
# Бинирование в цикле по атрибутам лонг-листа и параметрам модели
        
def binning_by_param_loop(attr_in_long_set,df_train,min_perc_total_list,stop_limit_list,min_perc_class_list,i):   
    listname = []
    dfname = pd.DataFrame()
    par_loop = 0 # номер цикла по переменной  
    # чтобы по категориальным переменным не бежал цикл по ненужному атрибуту
    if str(df_train[attr_in_long_set].dtype) == 'object':
        min_perc_total_list_fin = [0.05] # default
    else:
        min_perc_total_list_fin = min_perc_total_list
    # цикл по трем параметрам функции бинирования
    for stop_limit_attr in stop_limit_list:    
        for min_perc_class_attr in min_perc_class_list:       
            for min_perc_total_attr in min_perc_total_list_fin:                   
                # вызов функции бинирования                    
                attr_in_long_set_params = attr_in_long_set + '$_T_' + str(min_perc_total_attr) + '_C_' + str(min_perc_class_attr) + '_S_' + str(stop_limit_attr)
                par_loop = par_loop + 1
                print('4.1.' + str(i) + '.' + str(par_loop) + '. Бинирование переменной: ' \
                      + attr_in_long_set_params + ': ' + str(dtime.datetime.now()))   
                binning = woe_binning(df = df_train,                           # Исходная таблица с данными
                                      target_var = 'EVENT',                    # Название целевой функции
                                      pred_var = attr_in_long_set,             # Название бинируемой переменной
                                      min_perc_total = min_perc_total_attr,    # Кол-во изначальных бинов для количественных переменных
                                      min_perc_class = min_perc_class_attr,    # обязательно склеивает бин, если его WOE меньше данного значения
                                      stop_limit = stop_limit_attr,            # не склеивает бин дальше, если его WOE больше данного значения
                                      abbrev_fact_levels = 50,                 # аббревиатура для названия бинов при превышении их кол-ва?
                                      event_class = 'good')                    # good/bad - смысл целевой функции (1)?
                if len(binning) > 0:
                    binning['attr_name'] = attr_in_long_set_params
                    binning['attr_true_name'] = attr_in_long_set
                    binning['min_perc_total'] = min_perc_total_attr
                    binning['min_perc_class'] = min_perc_class_attr
                    binning['stop_limit'] = stop_limit_attr 
                    binning['par_loop'] = par_loop                
                    listname.append(binning)
                    dfname = pd.concat(listname,axis = 0)
                gc.collect() # принудительный вызов сборщика мусора
    return dfname

## binning_result_unification
# Приведение таблицы после функции бинирования в одинаковый вид
    
def binning_result_unification(binninglist):
    binningInfoLongInitial = binninglist.reset_index()
    binningInfoLongInitialList = list(binningInfoLongInitial)
    binningInfoLong = pd.DataFrame()
    
    for binningInfoLongInitialAttr in binningInfoLongInitialList:
        if binningInfoLongInitialAttr.lower() in ('attr_true_name',
                                                  'attr_name',
                                                  'good',
                                                  'bad',
                                                  'col_perc_a',
                                                  'col_perc_b',
                                                  'cutpoints_final',
                                                  'upper_cutpoints_final_dfrm',
                                                  'woe',
                                                  'iv_bins',
                                                  'iv_total_final',
                                                  'min_perc_total',
                                                  'min_perc_class',
                                                  'stop_limit',
                                                  'par_loop'):      
            binningInfoLong[binningInfoLongInitialAttr] = binningInfoLongInitial[binningInfoLongInitialAttr]
        if binningInfoLongInitialAttr.lower() in ('index','predictor_var_binned'):
            binningInfoLong['int_interval'] = binningInfoLongInitial[binningInfoLongInitialAttr].astype('str')
        if binningInfoLongInitialAttr.lower() in ('group_1','group_2'):
            binningInfoLong[binningInfoLongInitialAttr] = binningInfoLongInitial[binningInfoLongInitialAttr]
    
    if 'Group_1' not in binningInfoLongInitialList:
        binningInfoLong['Group_1'] = ''
    if 'Group_2' not in binningInfoLongInitialList:
        binningInfoLong['Group_2'] = None
    if 'cutpoints_final' not in binningInfoLongInitialList:
        binningInfoLong['cutpoints_final'] = ''
    if 'upper_cutpoints_final_dfrm' not in binningInfoLongInitialList:
        binningInfoLong['upper_cutpoints_final_dfrm'] = ''    
        
    binningInfoLong['cutpoints_final'] = binningInfoLong['cutpoints_final'].astype('str')
    binningInfoLong['upper_cutpoints_final_dfrm'] = binningInfoLong['upper_cutpoints_final_dfrm'].astype('str')
    return binningInfoLong

## binning_get_shortlist_after_iv
    
def binning_get_shortlist_after_iv(binningInfoLong,IV_left_cutoff,IV_right_cutoff):
    binningInfoShort = pd.DataFrame(binningInfoLong[['attr_true_name',
                                                 'attr_name',
                                                 'min_perc_total',
                                                 'min_perc_class',
                                                 'stop_limit',
                                                 'par_loop',
                                                 'iv_total_final']] \
                                .groupby(by = ['attr_true_name', 
                                               'attr_name',
                                               'min_perc_total',
                                               'min_perc_class',
                                               'stop_limit',
                                               'par_loop',
                                               'iv_total_final']) \
                                .count()
                            ).reset_index() \
                             .sort_values(by = 'iv_total_final', ascending = False) \
                             .reset_index() \
                             .drop('index', axis = 1)
                             
    binningInfoShort['iv_detail'] = ''
    binningInfoShort['iv_detail'] = binningInfoShort.apply(lambda x: get_IV_detail(x.iv_total_final, x.iv_detail, 'left', IV_left_cutoff), axis=1) 
    binningInfoShort['iv_detail'] = binningInfoShort.apply(lambda x: get_IV_detail(x.iv_total_final, x.iv_detail, 'right', IV_right_cutoff), axis=1)
    binningInfoShort['iv_check'] = binningInfoShort.apply(lambda x: get_detail_check(x.iv_detail), axis=1)
    return binningInfoShort

## binning_get_data_for_trend
    
def binning_get_data_for_trend(binningInfoLong):
    
    # общая единая унифицированная таблица
    
    binningInfoLongAll = pd.DataFrame()
    binningInfoLongAll['attr_true_name'] = binningInfoLong['attr_true_name']
    binningInfoLongAll['attr_name'] = binningInfoLong['attr_name']
    binningInfoLongAll['min_perc_total'] = binningInfoLong['min_perc_total']
    binningInfoLongAll['min_perc_class'] = binningInfoLong['min_perc_class']
    binningInfoLongAll['stop_limit'] = binningInfoLong['stop_limit']   
    binningInfoLongAll['par_loop'] = binningInfoLong['par_loop']
    binningInfoLongAll['bin_number'] = binningInfoLong.groupby(['attr_name'])['upper_cutpoints_final_dfrm'].cumcount()+1
    binningInfoLongAll['int_interval'] = binningInfoLong['int_interval']
    binningInfoLongAll['int_left_cut'] = binningInfoLong['cutpoints_final']
    binningInfoLongAll['int_right_cut'] = binningInfoLong['upper_cutpoints_final_dfrm']
    binningInfoLongAll['cat_attr_list'] = binningInfoLong['Group_1']
    binningInfoLongAll['cat_name'] = binningInfoLong['Group_2']
    binningInfoLongAll['quant_1'] = binningInfoLong['bad']
    binningInfoLongAll['quant_0'] = binningInfoLong['good']
    
    if len(binningInfoLong) > 0: 
        binningInfoLongAll['cat_attr_list'] = binningInfoLong.apply(lambda x: get_true_cat_attr_name(x.Group_1), axis=1) 
        binningInfoLongAll['woe'] = binningInfoLong.apply(lambda x: get_true_woe_value(x.woe), axis=1)
        
    
    # правки для интервальных переменных
    
    binningInfoLongInt = binningInfoLongAll[binningInfoLongAll.cat_name.isnull()]
    
    binningInfoLongIntModify = binningInfoLongInt
    binningInfoLongIntModify['interval_zpt'] = binningInfoLongIntModify['int_interval'].str.find(',')
    binningInfoLongIntModify['interval_right_scb'] = binningInfoLongIntModify['int_interval'].str.find(']')
      
    if len(binningInfoLongIntModify) > 0:  
        
        binningInfoLongIntModify['true_int_left_cut']  = binningInfoLongIntModify.apply(lambda x: get_true_int_cut(x.int_interval, \
                                                                                                                   x.interval_zpt, \
                                                                                                                   0, \
                                                                                                                   'left'), axis=1) 
    
        binningInfoLongIntModify['true_int_right_cut'] = binningInfoLongIntModify.apply(lambda x: get_true_int_cut(x.int_interval, \
                                                                                                                   x.interval_zpt, \
                                                                                                                   x.interval_right_scb, \
                                                                                                                   'right'), axis=1)
        binningInfoLongIntModify['int_left_cut']       = binningInfoLongIntModify['true_int_left_cut']
        binningInfoLongIntModify['int_right_cut']      = binningInfoLongIntModify['true_int_right_cut']
    
    # правки для категориальных переменных
    
    binningInfoLongCat = binningInfoLongAll[~binningInfoLongAll.cat_name.isnull()]

    binningInfoLongCatModify = binningInfoLongCat.groupby(by = ['attr_true_name',
                                                                'attr_name',
                                                                'min_perc_total',
                                                                'min_perc_class',
                                                                'stop_limit',
                                                                'par_loop',
                                                                'woe',
                                                                'cat_name'], as_index = False) \
                                                 .agg({'quant_1': 'mean',
                                                       'quant_0': 'mean',
                                                       'cat_attr_list' : 'sum'
                                                      })

    # сортировка категориальных бинов по возрастанию woe в рамках переменной
    
    binningInfoLongCatModify = binningInfoLongCatModify.sort_values(by = ['attr_name', 'woe'], ascending=[True, True])

    binningInfoLongCatModify['bin_number'] = binningInfoLongCatModify.groupby(['attr_name'])['woe'].cumcount() + 1    
    binningInfoLongCatModify['cat_attr_list'] = binningInfoLongCatModify['cat_attr_list'].astype('str')
    binningInfoLongCatModify['cat_attr_list'] = binningInfoLongCatModify['cat_attr_list'].str.slice(start = 0, stop = -2)
    
    # создание итоговой таблицы для расчета трендовости
    
    binningInfoLongFin = pd.concat([binningInfoLongIntModify, binningInfoLongCatModify], ignore_index = True) \
                           .sort_values(by = ['attr_name', 'bin_number'], ascending=[True, True])
    
    binningInfoLongFin = binningInfoLongFin[['attr_true_name',
                                             'attr_name', 
                                             'min_perc_total',
                                             'min_perc_class',
                                             'stop_limit',
                                             'par_loop',
                                             'woe',
                                             'bin_number',  
                                             'int_left_cut', 
                                             'int_right_cut',
                                             'cat_name',
                                             'cat_attr_list',
                                             'quant_0',
                                             'quant_1'
                                             ]]   
    return binningInfoLongFin

## pvalue_get_check
    
def pvalue_get_check(pValueInfo,df_train,PVALUE_cutoff,target_attr):
    
    pValueInfo['pvalue_detail'] = ''
    
    pValueBadAttrList = []
    pValueCheck = 1
    i = 0
    
    while pValueCheck > 0:    
        i = i + 1
        pValueInfoList = list(pValueInfo['attr_true_name'])
        
        if len(pValueBadAttrList) > 0:
            for badValue in pValueBadAttrList:
                pValueInfoList.remove(badValue)
        
        pValueDfTrain = df_train[pValueInfoList].select_dtypes(include = ['number'])
        pValueAttr = sm.add_constant(pValueDfTrain)
        pValueInfoList = list(pValueDfTrain)
        pValueTarget = df_train[target_attr]    
        pValueModel = sm.OLS(pValueTarget, pValueAttr)
        pValueResult = pValueModel.fit()   
        pValueResultValues = pd.DataFrame()
           
        for pValueAttrName in pValueInfoList:
            
            pValueInfo['pvalue_detail'] = pValueInfo.apply(lambda x: get_PValue_detail(x.attr_true_name, \
                                                                                       pValueAttrName, \
                                                                                       x.pvalue_detail, \
                                                                                       pValueResult.pvalues[pValueAttrName], \
                                                                                       PVALUE_cutoff, \
                                                                                       i), axis=1) 
            
            pValueResultValuesAttr = pd.DataFrame(columns =  ['attr_true_name', 'pValue'])
            pValueResultValuesAttr.loc[-1] = [pValueAttrName, pValueResult.pvalues[pValueAttrName]]
            pValueResultValues = pValueResultValues.append(pValueResultValuesAttr, ignore_index=True)
            
        if len(pValueResultValues) == 0:
            pValueResultValues = pd.DataFrame(columns =  ['attr_true_name', 'pValue'])
        
        pValueBadAttrListIteration = list(pValueResultValues.query('pValue > @PVALUE_cutoff').attr_true_name)
        pValueBadAttrList.extend(pValueBadAttrListIteration)
        pValueCheck = len(pValueBadAttrListIteration)    
        print('6.2.' + str(i) + '. Убрано ' + str(pValueCheck) +  ' переменных: ' + str(dtime.datetime.now()))
        
    pValueInfo['pvalue_check'] = pValueInfo.apply(lambda x: get_detail_check(x.pvalue_detail), axis=1)
    
    return pValueInfo

## pv_calculate
# Функция возвращает датафрейм с 2-мя столбцами - название переменной (поле Variable) и значение PV (поле P-value)
    
def pv_calculate(var_list, df_train, target_attr='EVENT'):
    
    #pv_list = list(var_batch[1])
    pv_df_train = df_train[var_list].select_dtypes(include = ['number'])
    pv_attr = sm.add_constant(pv_df_train)
    pv_target = df_train[target_attr]
    
    pv_Model = sm.OLS(pv_target, pv_attr)
    pv_Result = pv_Model.fit()
    df_batch = pd.DataFrame({'P-value' : pv_Result.pvalues}).reset_index().rename(columns={'index' : 'Variable'})
    df_batch['P-value'] = df_batch['P-value'].round(6)
    
    return df_batch[df_batch['Variable'] != 'const']

## get_pv_sequences

def get_pv_sequences(df_train, var_list, thresholds=[0.9, 0.8, 0.7, 0.6, 0.5, 0.4, 0.3, 0.2, 0.1], target_attr='EVENT'):
    
    dd_1 = pv_calculate(var_list, df_train, target_attr=target_attr)
    dd_2 = pd.DataFrame(columns=['attr_true_name', 'pvalue_detail'])
    dd_2['attr_true_name'] = dd_1['Variable']
    #print(dd_1)
    for k, threshold in enumerate(thresholds):
        #print('Порог ' + str(threshold))
        LL = list(dd_1.loc[dd_1['P-value'] >= threshold, 'Variable']) # список пер-х выше порога
        L = list(dd_1.loc[dd_1['P-value'] < threshold, 'Variable'])  # список пер-х ниже порога
        
        i = 1
        while len(LL) > 0:
        
            dd_2.loc[dd_2['attr_true_name'].isin(LL), 'pvalue_detail'] = '3.' + str(k+1) + '. PValue превышает порог ' + str(threshold) + '; итерация ' + str(k+1) + '.' + str(i)
            
            dd_1 = pv_calculate(L, df_train, target_attr=target_attr)
            #print('Цикл ' + str(i))
            #print(dd_1)
            LL = list(dd_1.loc[dd_1['P-value'] >= threshold, 'Variable']) # список пер-х выше порога
            L = list(dd_1.loc[dd_1['P-value'] < threshold, 'Variable'])  # список пер-х ниже порога
            
            i += 1
    
    dd_2[['pvalue_detail']] = dd_2[['pvalue_detail']].fillna('')    
    dd_2['pvalue_check'] = dd_2.apply(lambda x: get_detail_check(x.pvalue_detail), axis=1)
    dd_2 = dd_2.set_index('attr_true_name', drop = True)
    return dd_2

## get_gini_gap

def get_gini_gap(gini_train, gini_test):
    return abs(gini_train - gini_test)

## get_gini_gap_flg

def get_gini_gap_flg(gini_train, gini_test):
    if (abs(gini_train - gini_test) <= 0.02) and (gini_test > 0):
        return 1
    else:
        return 0
    
## cramers_corrected_stat

def cramers_corrected_stat(confusion_matrix):
    """ calculate Cramers V statistic for categorial-categorial association.
        uses correction from Bergsma and Wicher, 
        Journal of the Korean Statistical Society 42 (2013): 323-328
        https://ru.stackoverflow.com/questions/927487
    """
    chi2 = ss.chi2_contingency(confusion_matrix)[0]
    n = confusion_matrix.sum().sum()
    phi2 = chi2/n
    r,k = confusion_matrix.shape
    phi2corr = max(0, phi2 - ((k-1)*(r-1))/(n-1))    
    rcorr = r - ((r-1)**2)/(n-1)
    kcorr = k - ((k-1)**2)/(n-1)
    return np.sqrt(phi2corr / min( (kcorr-1), (rcorr-1)))

# Корреляция Крамера для категориальных переменных. Возвращает матрицу корреляции
# Список пер-х, проверяемых на корреляцию
# Датафрем, содержащий EVENT и переменные, проверяемые на корреляцию

def cat_variable_corr(var_list, df, CORREL_cutoff=0.50):
    
    if len(var_list) >= 2:
    
        correlInfoList = var_list
        correlDataCat = df[correlInfoList]#.select_dtypes(include = ['object', 'category'])
        correlDataCatList = list(correlDataCat)
        correlBadAttrListCat = []

        ## Проверка категориальных переменных

        if correlDataCat.shape[1] > 1:
            for i in correlDataCatList:
                for j in correlDataCatList:
                    if i != j \
                        and i not in correlBadAttrListCat \
                        and j not in correlBadAttrListCat:
                            confusion_mx = pd.crosstab(correlDataCat[i],correlDataCat[j])
                            corrCatValue = cramers_corrected_stat(confusion_mx)
                            if abs(corrCatValue) > CORREL_cutoff:
                                correlBadAttrListCat.append(j)

        var_list_ = set(var_list).difference(set(correlBadAttrListCat))
        var_list_ = list(var_list_) # список переменных, отобравшихся по корр.

        corrDataMtrxCat = pd.DataFrame(var_list_, columns=['attr'])
        for attr_name in var_list_:
            corrDataMtrxCat[attr_name] = 0.0
        corrDataMtrxCat = corrDataMtrxCat.set_index('attr', drop=True)

        for i in var_list_:
            for j in var_list_:
                confusion_mx = pd.crosstab(df[i], df[j])
                corrCatValue = cramers_corrected_stat(confusion_mx)
                corrDataMtrxCat.at[i,j] = corrCatValue

        return corrDataMtrxCat.loc[var_list, var_list]
    
    else:
        print('!!! Для построения корреляционных соотношений необходимы 2 или более пер-х !!!')
            
# Добавляем матрицу корреляции в отчет
# corrmtrx - корр. мат. (в индексе - название пер-х)

def corrmtrx_to_excel(corrmtrx, wb):
    
    if len(corrmtrx) >= 2:
    
        sheet = wb.create_sheet('CorrMtrx')
        
        sheet.cell(row=1, column=1).value = 'Variables'
        sheet.cell(row=1, column=1).font = Font(bold=True)
        sheet.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')

        col = list(corrmtrx)
        
        max_length = 0
        for k, j in enumerate(col):
            
            if len(j) >  max_length:
                 max_length = len(j)

            n = 1
            sheet.cell(row=1, column=k+2).value = j
            sheet.cell(row=1, column=k+2).border = thin_border_top_bot_l_r
            sheet.cell(row=k+2, column=1).value = j
            sheet.cell(row=k+2, column=1).border = thin_border_top_bot_l_r

            for i, row in corrmtrx.iterrows():
                n += 1
                sheet.cell(row=n, column=k+2).value = round(row[j] * 100, 2)
                sheet.cell(row=n, column=k+2).border = thin_border_top_bot_l_r
                
        sheet.column_dimensions['A'].width = (max_length + 2) * 1.2
        
        rule = ColorScaleRule(start_type='min',
                                      #start_value=10,
                                      start_color='FF63BE7B',
                                      mid_type='percentile',
                                      mid_value=50,
                                      mid_color='FFFFEB84', 
                                      end_type='max',
                                      #end_value=90,
                                      end_color='FFF8696B')
        
        

        #range_string = 'B2:' + get_column_letter(len(col)+1) + str(len(col)+1)
        
        # Получаем набор полей без центральной диагонали для раскраски
        
        corrMtrxRange = range(2,2+len(col))
        range_string = ''
        
        for i in corrMtrxRange: 
            
            minabs = None  
            beforecol = None   
            aftercol = None   
            maxabs = None  
            first_interval = None   
            second_interval = None  
            
            for j in corrMtrxRange:
                if j != i:
                    if j < i and minabs is None :
                        minabs = j
                    if j < i:
                        beforecol = j
                    if j > i and aftercol is None :
                        aftercol = j
                    if j > i:
                        maxabs = j
                        
            if minabs is None or beforecol is None:
                first_interval = ''
            elif minabs == beforecol:
                first_interval = get_column_letter(i) + str(minabs)
            else:
                first_interval = get_column_letter(i) + str(minabs) + ':' + get_column_letter(i) + str(beforecol)
                
            if aftercol is None or maxabs is None:
                second_interval = ''
            elif aftercol == maxabs:
                second_interval = get_column_letter(i) + str(aftercol)
            else:
                second_interval = get_column_letter(i) + str(aftercol) + ':' + get_column_letter(i) + str(maxabs)
            
            if first_interval == '':
                range_string = range_string + second_interval
            elif first_interval != '' and second_interval != '':
                range_string = range_string + ' ' + first_interval + ' ' + second_interval
            else:
                range_string = range_string + ' ' + first_interval
        
        # Применяем условное форматирование
        
        sheet.conditional_formatting.add(range_string, rule)
        
    else:
        
        sheet = wb.create_sheet('CorrMtrx')
        
        sheet['B2'] = 'Недостаточное кол-во пер-х для построения корреляционной матрицы'
        sheet.column_dimensions['B'].width = 70

## append_df_to_excel

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    @param filename: File path or existing ExcelWriter
                     (Example: '/path/to/file.xlsx')
    @param df: DataFrame to save to workbook
    @param sheet_name: Name of sheet which will contain DataFrame.
                       (default: 'Sheet1')
    @param startrow: upper left cell row to dump data frame.
                     Per default (startrow=None) calculate the last row
                     in the existing DF and write to the next row...
    @param truncate_sheet: truncate (remove and recreate) [sheet_name]
                           before writing DataFrame to Excel file
    @param to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()`
                            [can be a dictionary]
    @return: None

    Usage examples:

    >>> append_df_to_excel('d:/temp/test.xlsx', df)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, header=None, index=False)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
                           index=False)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2', 
                           index=False, startrow=25)

    (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
    """
    # Excel file doesn't exist - saving and exiting
    if not os.path.isfile(filename):
        df.to_excel(
            filename,
            sheet_name=sheet_name, 
            startrow=startrow if startrow is not None else 0, 
            **to_excel_kwargs)
        return
    
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')

    # try to open an existing workbook
    writer.book = load_workbook(filename)
    
    # get the last row in the existing Excel sheet
    # if it was not specified explicitly
    if startrow is None and sheet_name in writer.book.sheetnames:
        startrow = writer.book[sheet_name].max_row

    # truncate sheet
    if truncate_sheet and sheet_name in writer.book.sheetnames:
        # index of [sheet_name] sheet
        idx = writer.book.sheetnames.index(sheet_name)
        # remove [sheet_name]
        writer.book.remove(writer.book.worksheets[idx])
        # create an empty sheet [sheet_name] using old index
        writer.book.create_sheet(sheet_name, idx)
    
    # copy existing sheets
    writer.sheets = {ws.title:ws for ws in writer.book.worksheets}

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()
    
## Получение значения бинов переменных

def bin_df(df, bin_Results, short_lst, var_bin_type = 'woe'):
    
    # Если  var_bin_type = 'bin_number', то переводим значения переменных в номера бинов
    # Если  var_bin_type = 'woe', то переводим значения переменных в WOE
    
    # Обрабатываем интервальные переменные
    
    df_bin =  df.loc[:, ['EVENT']]

    var_name_all = list(bin_Results[bin_Results['int_right_cut'].notna()]['attr_name'].unique())
    var_name = [i for i in var_name_all if i in short_lst]

    for variable in var_name:

        var_test = variable[0:variable.find('$')]

        df_var = bin_Results.loc[bin_Results['attr_name'] == variable, ['bin_number', 'int_right_cut', 'woe']].sort_values(by='bin_number', ascending=True)
        df_var['int_right_cut'] = df_var['int_right_cut'].astype('float64')

        df_1 = df.loc[:, ['EVENT', var_test]]

        condlist = []
        choicelist = []

        for bin_num, right_cut in zip(df_var[var_bin_type], df_var['int_right_cut']):
            condlist.append(df_1[var_test] <= right_cut)
            choicelist.append(bin_num)

        df_bin[var_test] = np.select(condlist, choicelist)


    # Обрабатываем категориальные переменные

    var_cat_name_all = list(bin_Results[bin_Results['int_right_cut'].isnull()]['attr_name'].unique())
    var_cat_name = [i for i in var_cat_name_all if i in short_lst]

    for variable in var_cat_name:

        var_test = variable[0:variable.find('$')]

        df_var = bin_Results.loc[bin_Results['attr_name'] == variable, ['bin_number', 'cat_attr_list', 'woe']].sort_values(by='bin_number', ascending=True)

        df_1 = df.loc[:, ['EVENT', var_test]]

        condlist = []
        choicelist = []

        for bin_num, cat_attr in zip(df_var[var_bin_type], df_var['cat_attr_list']):
            condlist.append(df_1[var_test].isin(cat_attr.split(', ')))
            choicelist.append(bin_num)

        df_bin[var_test] = np.select(condlist, choicelist)
    
    return df_bin

## run_randomsearch

def run_randomsearch(X, 
                     y, 
                     clf, 
                     param, 
                     cv = 5, 
                     n_iter_search = 20,
                     random_state = 1):

    random_search = RandomizedSearchCV(estimator = clf, 
                                       param_distributions = param,
                                       n_iter = n_iter_search,
                                       cv = cv,
                                       random_state = random_state,
                                       scoring = 'roc_auc')    
    random_search.fit(X, y)
    return random_search

## Расчет ROC_AUC

def roc_auc(df, df_, wb3, n_bucket=10):
    
    # подготовка 1-го датафрейма
    rn = np.floor(np.array(range(1, len(df) + 1)) / (len(df) + 1) * n_bucket) + 1

    df = df.sort_values(by='SCORE', ascending=True)
    df['group'] = pd.Series(rn[rn <= n_bucket], index=df.index)

    df_2 = df.groupby(by='group').aggregate({'SCORE': 'count', 'EVENT': 'sum'}) \
                                 .rename(columns={'SCORE': 'OBS,#', 'EVENT': 'Goods,#'}) \
                                 .reset_index().sort_values(by='group', ascending=True)

    df_2['Bads,#'] = df_2['OBS,#'] - df_2['Goods,#']
    
    
    # подготовка 2-го датафрейма
    rn_ = np.floor(np.array(range(1, len(df_) + 1)) / (len(df_) + 1) * n_bucket) + 1

    df_ = df_.sort_values(by='SCORE', ascending=True)
    df_['group'] = pd.Series(rn_[rn_ <= n_bucket], index=df_.index)

    df_2_ = df_.groupby(by='group').aggregate({'SCORE': 'count', 'EVENT': 'sum'}) \
                                   .rename(columns={'SCORE': 'OBS,#', 'EVENT': 'Goods,#'}) \
                                   .reset_index().sort_values(by='group', ascending=True)

    df_2_['Bads,#'] = df_2_['OBS,#'] - df_2_['Goods,#']

    #sheet = wb3.active
    #sheet.title = 'ROC_AUC'
    sheet = wb3.create_sheet('ROC_AUC')
    
    chart1 = ScatterChart()
    chart1.title = "ROC curve"
    chart1.x_axis.title = 'Goods,%++'
    chart1.y_axis.title = 'Bads,%++'
    chart1.legend = None
    chart1.height = 10 # default is 7.5
    chart1.width = 15 # default is 15

    chart1.x_axis.scaling.min = 0
    chart1.y_axis.scaling.min = 0
    chart1.x_axis.scaling.max = 100
    chart1.y_axis.scaling.max = 100
    
    sheet['AF2'] = 0
    sheet['AF3'] = 100
    sheet['AG2'] = 0
    sheet['AG3'] = 100
    
    sheet['L'+str(n_bucket+6)] = 'AUC'
    sheet['L'+str(n_bucket+6)].font = Font(bold=True)
    sheet['L'+str(n_bucket+6)].border = thin_border_top_bot_r
    sheet['L'+str(n_bucket+6)].alignment = Alignment(horizontal='center', vertical='center')
    sheet['L'+str(n_bucket+7)].border = thin_border_top_bot_r
    sheet['L'+str(n_bucket+8)].border = thin_border_top_bot_r
    sheet['M'+str(n_bucket+6)] = 'GINI'
    sheet['M'+str(n_bucket+6)].font = Font(bold=True)
    sheet['M'+str(n_bucket+6)].border = thin_border_top_bot_r
    sheet['M'+str(n_bucket+6)].alignment = Alignment(horizontal='center', vertical='center')
    sheet['M'+str(n_bucket+7)].border = thin_border_top_bot_r
    sheet['M'+str(n_bucket+8)].border = thin_border_top_bot_r
    sheet['N'+str(n_bucket+6)] = 'KSI'
    sheet['N'+str(n_bucket+6)].font = Font(bold=True)
    sheet['N'+str(n_bucket+6)].border = thin_border_top_bot_r
    sheet['N'+str(n_bucket+6)].alignment = Alignment(horizontal='center', vertical='center')
    sheet['N'+str(n_bucket+7)].border = thin_border_top_bot_r
    sheet['N'+str(n_bucket+8)].border = thin_border_top_bot_r
    
    sheet['O'+str(n_bucket+6)] = 'All'
    sheet['O'+str(n_bucket+6)].font = Font(bold=True)
    sheet['O'+str(n_bucket+6)].border = thin_border_top_bot_r
    sheet['O'+str(n_bucket+6)].alignment = Alignment(horizontal='center', vertical='center')
    sheet['O'+str(n_bucket+7)].border = thin_border_top_bot_r
    sheet['O'+str(n_bucket+8)].border = thin_border_top_bot_r
    
    sheet['P'+str(n_bucket+6)] = 'Good'
    sheet['P'+str(n_bucket+6)].font = Font(bold=True)
    sheet['P'+str(n_bucket+6)].border = thin_border_top_bot_r
    sheet['P'+str(n_bucket+6)].alignment = Alignment(horizontal='center', vertical='center')
    sheet['P'+str(n_bucket+7)].border = thin_border_top_bot_r
    sheet['P'+str(n_bucket+8)].border = thin_border_top_bot_r
    
    sheet['Q'+str(n_bucket+6)] = 'Bad'
    sheet['Q'+str(n_bucket+6)].font = Font(bold=True)
    sheet['Q'+str(n_bucket+6)].border = thin_border_top_bot_r
    sheet['Q'+str(n_bucket+6)].alignment = Alignment(horizontal='center', vertical='center')
    sheet['Q'+str(n_bucket+7)].border = thin_border_top_bot_r
    sheet['Q'+str(n_bucket+8)].border = thin_border_top_bot_r
    
    sheet['R'+str(n_bucket+6)] = 'Good/Bad,%'
    sheet['R'+str(n_bucket+6)].font = Font(bold=True)
    sheet['R'+str(n_bucket+6)].border = thin_border_top_bot_r
    sheet['R'+str(n_bucket+6)].alignment = Alignment(horizontal='center', vertical='center')
    sheet['R'+str(n_bucket+7)].border = thin_border_top_bot_r
    sheet['R'+str(n_bucket+8)].border = thin_border_top_bot_r
    
    sheet['K'+str(n_bucket+6)].border = thin_border_top_bot_l_r
    sheet['K'+str(n_bucket+7)] = 'train'
    sheet['K'+str(n_bucket+7)].font = Font(bold=True)
    sheet['K'+str(n_bucket+7)].border = thin_border_l_r
    sheet['K'+str(n_bucket+7)].alignment = Alignment(horizontal='center', vertical='center')
    sheet['K'+str(n_bucket+8)] = 'test'
    sheet['K'+str(n_bucket+8)].font = Font(bold=True)
    sheet['K'+str(n_bucket+8)].border = thin_border_top_bot_l_r
    sheet['K'+str(n_bucket+8)].alignment = Alignment(horizontal='center', vertical='center')
    
    m = 1
    
    sheet.cell(row=1, column=2).value = 'train'
    sheet.cell(row=1, column=2).font = Font(bold=True)
    sheet.cell(row=1, column=2).border = thin_border_top_bot_l_r
    sheet.cell(row=1, column=2).fill = PatternFill(start_color='00969696', fill_type = "solid")
    sheet.cell(row=1, column=2).alignment = Alignment(horizontal='center', vertical='center')
    sheet.cell(row=1, column=16).value = 'test'
    sheet.cell(row=1, column=16).font = Font(bold=True)
    sheet.cell(row=1, column=16).border = thin_border_top_bot_l_r
    sheet.cell(row=1, column=16).fill = PatternFill(start_color='00969696', fill_type = "solid")
    sheet.cell(row=1, column=16).alignment = Alignment(horizontal='center', vertical='center')
    
    sheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=13)
    sheet.merge_cells(start_row=1, start_column=16, end_row=1, end_column=27)

    for i, df in zip((2, 16), (df_2, df_2_)):
        sheet.cell(row=2, column=i).value = 'Bucket'
        sheet.cell(row=2, column=i).font = Font(bold=True)
        sheet.cell(row=2, column=i).border = thin_border_top_bot_l_r
        sheet.cell(row=2, column=i).alignment = Alignment(horizontal='center', vertical='center')
        sheet.cell(row=2, column=i+1).value = 'OBS,#'
        sheet.cell(row=2, column=i+1).font = Font(bold=True)
        sheet.cell(row=2, column=i+1).border = thin_border_top_bot_r
        sheet.cell(row=2, column=i+1).alignment = Alignment(horizontal='center', vertical='center')
        sheet.cell(row=2, column=i+2).value = 'Bads,%'
        sheet.cell(row=2, column=i+2).font = Font(bold=True)
        sheet.cell(row=2, column=i+2).border = thin_border_top_bot_r
        sheet.cell(row=2, column=i+2).alignment = Alignment(horizontal='center', vertical='center')
        sheet.cell(row=2, column=i+3).value = 'p(Bad),%'
        sheet.cell(row=2, column=i+3).font = Font(bold=True) 
        sheet.cell(row=2, column=i+3).border = thin_border_top_bot_r
        sheet.cell(row=2, column=i+3).alignment = Alignment(horizontal='center', vertical='center')
        sheet.cell(row=2, column=i+4).value = 'Bads,#'
        sheet.cell(row=2, column=i+4).font = Font(bold=True) 
        sheet.cell(row=2, column=i+4).border = thin_border_top_bot_r
        sheet.cell(row=2, column=i+4).alignment = Alignment(horizontal='center', vertical='center')
        sheet.cell(row=2, column=i+5).value = 'Bads,#++'
        sheet.cell(row=2, column=i+5).font = Font(bold=True) 
        sheet.cell(row=2, column=i+5).border = thin_border_top_bot_r
        sheet.cell(row=2, column=i+5).alignment = Alignment(horizontal='center', vertical='center')
        sheet.cell(row=2, column=i+6).value = 'Goods,#'
        sheet.cell(row=2, column=i+6).font = Font(bold=True) 
        sheet.cell(row=2, column=i+6).border = thin_border_top_bot_r
        sheet.cell(row=2, column=i+6).alignment = Alignment(horizontal='center', vertical='center')
        sheet.cell(row=2, column=i+7).value = 'Goods,#++'
        sheet.cell(row=2, column=i+7).font = Font(bold=True) 
        sheet.cell(row=2, column=i+7).border = thin_border_top_bot_r
        sheet.cell(row=2, column=i+7).alignment = Alignment(horizontal='center', vertical='center')
        sheet.cell(row=2, column=i+8).value = 'Bads,%++'
        sheet.cell(row=2, column=i+8).font = Font(bold=True) 
        sheet.cell(row=2, column=i+8).border = thin_border_top_bot_r
        sheet.cell(row=2, column=i+8).alignment = Alignment(horizontal='center', vertical='center')
        sheet.cell(row=2, column=i+9).value = 'Goods,%++'
        sheet.cell(row=2, column=i+9).font = Font(bold=True) 
        sheet.cell(row=2, column=i+9).border = thin_border_top_bot_r
        sheet.cell(row=2, column=i+9).alignment = Alignment(horizontal='center', vertical='center')
        sheet.cell(row=2, column=i+10).value = 'AROCi,%'
        sheet.cell(row=2, column=i+10).font = Font(bold=True) 
        sheet.cell(row=2, column=i+10).border = thin_border_top_bot_r
        sheet.cell(row=2, column=i+10).alignment = Alignment(horizontal='center', vertical='center')
        sheet.cell(row=2, column=i+11).value = 'Ksi,%'
        sheet.cell(row=2, column=i+11).font = Font(bold=True) 
        sheet.cell(row=2, column=i+11).border = thin_border_top_bot_r
        sheet.cell(row=2, column=i+11).alignment = Alignment(horizontal='center', vertical='center')

        sheet.cell(row=3, column=i+8).value = 0
        sheet.cell(row=3, column=i+9).value = 0
        
        sheet.cell(row=3, column=i).border = thin_border_top_bot_l
        sheet.cell(row=3, column=i).fill = PatternFill(start_color='00C0C0C0', fill_type = "solid")
        sheet.cell(row=3, column=i+1).border = thin_border_bot
        sheet.cell(row=3, column=i+1).fill = PatternFill(start_color='00C0C0C0', fill_type = "solid")
        sheet.cell(row=3, column=i+2).border = thin_border_bot
        sheet.cell(row=3, column=i+2).fill = PatternFill(start_color='00C0C0C0', fill_type = "solid")
        sheet.cell(row=3, column=i+3).border = thin_border_bot
        sheet.cell(row=3, column=i+3).fill = PatternFill(start_color='00C0C0C0', fill_type = "solid")
        sheet.cell(row=3, column=i+4).border = thin_border_bot
        sheet.cell(row=3, column=i+4).fill = PatternFill(start_color='00C0C0C0', fill_type = "solid")
        sheet.cell(row=3, column=i+5).border = thin_border_bot
        sheet.cell(row=3, column=i+5).fill = PatternFill(start_color='00C0C0C0', fill_type = "solid")
        sheet.cell(row=3, column=i+6).border = thin_border_bot
        sheet.cell(row=3, column=i+6).fill = PatternFill(start_color='00C0C0C0', fill_type = "solid")
        sheet.cell(row=3, column=i+7).border = thin_border_bot_r
        sheet.cell(row=3, column=i+7).fill = PatternFill(start_color='00C0C0C0', fill_type = "solid")
        sheet.cell(row=3, column=i+8).border = thin_border_bot_r
        sheet.cell(row=3, column=i+8).fill = PatternFill(start_color='00C0C0C0', fill_type = "solid")
        sheet.cell(row=3, column=i+9).border = thin_border_bot_r
        sheet.cell(row=3, column=i+9).fill = PatternFill(start_color='00C0C0C0', fill_type = "solid")
        sheet.cell(row=3, column=i+10).border = thin_border_bot
        sheet.cell(row=3, column=i+10).fill = PatternFill(start_color='00C0C0C0', fill_type = "solid")
        sheet.cell(row=3, column=i+11).border = thin_border_bot_r
        sheet.cell(row=3, column=i+11).fill = PatternFill(start_color='00C0C0C0', fill_type = "solid")

        Goods_acc = 0
        Bads_acc = 0

        Goods_acc_sum = df['Goods,#'].sum()
        Bads_acc_sum = df['Bads,#'].sum()

        Goods_acc_inter = 0
        Bads_acc_inter = 0

        AUC = 0
        KSI = []

        for k, row in df.iterrows():

            Goods_acc += row['Goods,#']
            Bads_acc += row['Bads,#']

            Goods_acc_per = round(Goods_acc * 100 / Goods_acc_sum, 2)
            Bads_acc_per = round(Bads_acc * 100 / Bads_acc_sum, 2)

            sheet.cell(row=k+1 + 3, column=i).value = k + 1
            sheet.cell(row=k+1 + 3, column=i).border = thin_border_l_r
            sheet.cell(row=k+1 + 3, column=i+1).value = row['OBS,#']
            sheet.cell(row=k+1 + 3, column=i+1).border = thin_border_l_r
            #sheet.cell(row=k+1 + 3, column=i+2).value = round(row['Bads,#'] * 100 / row['OBS,#'], 2)
            #sheet.cell(row=k+1 + 3, column=i+3).value = ['p(Bad),%']
            sheet.cell(row=k+1 + 3, column=i+4).value = row['Bads,#']
            sheet.cell(row=k+1 + 3, column=i+4).border = thin_border_l_r
            sheet.cell(row=k+1 + 3, column=i+5).value = Bads_acc
            sheet.cell(row=k+1 + 3, column=i+5).border = thin_border_l_r
            sheet.cell(row=k+1 + 3, column=i+6).value = row['Goods,#']
            sheet.cell(row=k+1 + 3, column=i+6).border = thin_border_l_r
            sheet.cell(row=k+1 + 3, column=i+7).value = Goods_acc
            sheet.cell(row=k+1 + 3, column=i+7).border = thin_border_l_r
            sheet.cell(row=k+1 + 3, column=i+8).value = Bads_acc_per
            sheet.cell(row=k+1 + 3, column=i+8).border = thin_border_l_r
            sheet.cell(row=k+1 + 3, column=i+9).value = Goods_acc_per
            sheet.cell(row=k+1 + 3, column=i+9).border = thin_border_l_r
            sheet.cell(row=k+1 + 3, column=i+10).value = round((Goods_acc_per - Goods_acc_inter) * (Bads_acc_per + Bads_acc_inter) * 100 / 2 / 100 / 100, 2)
            sheet.cell(row=k+1 + 3, column=i+10).border = thin_border_l_r
            sheet.cell(row=k+1 + 3, column=i+11).value = Bads_acc_per - Goods_acc_per
            sheet.cell(row=k+1 + 3, column=i+11).border = thin_border_l_r

            AUC += round((Goods_acc_per - Goods_acc_inter) * (Bads_acc_per + Bads_acc_inter) * 100 / 2 / 100 / 100, 2)
            KSI.append(Bads_acc_per - Goods_acc_per)

            Goods_acc_inter = Goods_acc_per
            Bads_acc_inter = Bads_acc_per

            #sheet.cell(row=k+1 + 2, column=i+11).value = 
            
        for j in range(0, 11 + 1):
            sheet.cell(row=n_bucket+3+1, column=i + j).border = thin_border_top

        sheet['L' + str(n_bucket+6+m)] = AUC
        sheet['M' + str(n_bucket+6+m)] = 2 * AUC - 100
        sheet['N' + str(n_bucket+6+m)] = max(KSI)
        sheet['O' + str(n_bucket+6+m)] = df['OBS,#'].sum()
        sheet['P' + str(n_bucket+6+m)] = df['Goods,#'].sum()
        sheet['Q' + str(n_bucket+6+m)] = df['Bads,#'].sum()
        sheet['R' + str(n_bucket+6+m)] = round(df['Goods,#'].sum() * 100 / df['Bads,#'].sum(), 4)
        
        m += 1
            
        x = Reference(sheet, min_col=i+9, min_row=3, max_row=n_bucket + 3)
        y = Reference(sheet, min_col=i+8, min_row=3, max_row=n_bucket + 3)
        s = Series(y, x)
        s.smooth = True
        chart1.append(s)
        
        chart1.smooth = True
        
    sheet['M' + str(n_bucket+6+m)] = sheet['M' + str(n_bucket+6+m-2)].value - sheet['M' + str(n_bucket+6+m-1)].value
    sheet['N' + str(n_bucket+6+m)] = sheet['N' + str(n_bucket+6+m-2)].value - sheet['N' + str(n_bucket+6+m-1)].value
        
    x1 = Reference(sheet, min_col=32, min_row=2, max_row=3)
    y1 = Reference(sheet, min_col=33, min_row=2, max_row=3)
    s1 = Series(y1, x1)
    chart1.append(s1)

    sheet.add_chart(chart1, "B" + str(n_bucket + 6))
    
## Predict function

def predict(df, df_, wb2, n_bucket=20):
    
    s = 1
    
    d_ssi = {}
    
    sheet2 = wb2.create_sheet('Predict')
    
    for data_n, df in zip(('train', 'test'), (df, df_)):
    
        a = np.min(df['SCORE'])
        b = np.max(df['SCORE'])

        h = round((b - a) / n_bucket, 6)
        
        if h == 0.0:
            print('Невозможно просчитать Predict!')
            break
            
        #np.linspace(0, 5, 5)
        buckets = np.linspace(0, b, n_bucket+1)
        #np.append(np.arange(0, b, h), b)

        condlist = []
        choicelist = []

        for bucket, right_bound in list(enumerate(buckets))[1:]:
            condlist.append(df['SCORE'] <= right_bound)
            choicelist.append(bucket)

        df['bucket'] = np.select(condlist, choicelist)

        df_1 = df.groupby(by='bucket').aggregate({'SCORE': 'count', 'EVENT': 'sum'}) \
                                      .rename(columns={'SCORE': 'CNT', 'EVENT': 'GOODs'}) \
                                      .reset_index().sort_values(by='bucket', ascending=True)

        for i in choicelist:
            if df_1[df_1['bucket'] == i].empty:
                df_1 = df_1.append({'bucket': i, 'CNT': 0, 'GOODs': 0}, ignore_index=True)

        df_1 = df_1.sort_values(by='bucket', ascending=True)
        df_1['BADs'] = df_1['CNT'] - df_1['GOODs']

        cnt_sum = df_1['CNT'].sum()
        good_sum = df_1['GOODs'].sum()
        bad_sum = df_1['BADs'].sum()

        df_1['CNT, %'] = round(df_1['CNT'] * 100 / cnt_sum, 2)
        df_1['GOODs, %'] = round(df_1['GOODs'] * 100 / good_sum, 2)
        df_1['BADs, %'] = round(df_1['BADs'] * 100 / bad_sum, 2)
        df_1['Good/OBS, %'] = round(df_1['GOODs, %'] * 100 / (df_1['GOODs, %'] + df_1['BADs, %']), 2)
        df_1['Good/OBS, %'] = df_1['Good/OBS, %'].fillna(0)

        df_1['right_bound'] = pd.Series(buckets[1:], index=df_1.index)
        
        d_ssi.setdefault(data_n, df_1['CNT'] / cnt_sum)

        #wb2 = openpyxl.Workbook()
        #sheet2 = wb2.active
        #sheet2.title = 'Predict'
        
        sheet2.column_dimensions['B'].width = 15
        sheet2.column_dimensions['C'].width = 15
        sheet2.column_dimensions['L'].width = 21

        sheet2['A' + str(s)] = data_n
        sheet2['A' + str(s)].font = Font(bold=True)
        sheet2['A' + str(s)].border = thin_border_top_bot_l_r
        sheet2['A' + str(s)].alignment = Alignment(horizontal='center', vertical='center')
        sheet2['A' + str(s+1)] = 'Bucket'
        sheet2['A' + str(s+1)].font = Font(bold=True)
        sheet2['A' + str(s+1)].border = thin_border_top_bot_r
        sheet2['A' + str(s+1)].alignment = Alignment(horizontal='center', vertical='center')
        sheet2['B' + str(s+1)] = 'begin,%'
        sheet2['B' + str(s+1)].font = Font(bold=True)
        sheet2['B' + str(s+1)].border = thin_border_top_bot_r
        sheet2['B' + str(s+1)].alignment = Alignment(horizontal='center', vertical='center')
        sheet2['C' + str(s+1)] = 'end,%'
        sheet2['C' + str(s+1)].font = Font(bold=True)
        sheet2['C' + str(s+1)].border = thin_border_top_bot_r
        sheet2['C' + str(s+1)].alignment = Alignment(horizontal='center', vertical='center')
        sheet2['D' + str(s+1)] = 'OBS,#'
        sheet2['D' + str(s+1)].font = Font(bold=True)
        sheet2['D' + str(s+1)].border = thin_border_top_bot_r
        sheet2['D' + str(s+1)].alignment = Alignment(horizontal='center', vertical='center')
        sheet2['E' + str(s+1)] = 'Goods,#'
        sheet2['E' + str(s+1)].font = Font(bold=True)
        sheet2['E' + str(s+1)].border = thin_border_top_bot_r
        sheet2['E' + str(s+1)].alignment = Alignment(horizontal='center', vertical='center')
        sheet2['F' + str(s+1)] = 'Bads,#'
        sheet2['F' + str(s+1)].font = Font(bold=True)
        sheet2['F' + str(s+1)].border = thin_border_top_bot_r
        sheet2['F' + str(s+1)].alignment = Alignment(horizontal='center', vertical='center')
        sheet2['G' + str(s+1)] = 'OBS,%'
        sheet2['G' + str(s+1)].font = Font(bold=True)
        sheet2['G' + str(s+1)].border = thin_border_top_bot_r
        sheet2['G' + str(s+1)].alignment = Alignment(horizontal='center', vertical='center')
        sheet2['H' + str(s+1)] = 'Goods,%'
        sheet2['H' + str(s+1)].font = Font(bold=True)
        sheet2['H' + str(s+1)].border = thin_border_top_bot_r
        sheet2['H' + str(s+1)].alignment = Alignment(horizontal='center', vertical='center')
        sheet2['I' + str(s+1)] = 'Bads,%'
        sheet2['I' + str(s+1)].font = Font(bold=True)
        sheet2['I' + str(s+1)].border = thin_border_top_bot_r
        sheet2['I' + str(s+1)].alignment = Alignment(horizontal='center', vertical='center')
        sheet2['J' + str(s+1)] = 'Good/OBS,%'
        sheet2['J' + str(s+1)].font = Font(bold=True)
        sheet2['J' + str(s+1)].border = thin_border_top_bot_r
        sheet2['J' + str(s+1)].alignment = Alignment(horizontal='center', vertical='center')
        sheet2['L' + str(s+1)] = 'Проникновение ЦФ,%'
        sheet2['L' + str(s+1)].border = thin_border_top_bot_l_r
        sheet2['L' + str(s+1)].font = Font(bold=True)
        sheet2['L' + str(s+1)].alignment = Alignment(horizontal='center', vertical='center')

        sheet2.merge_cells(start_row=s, start_column=1, end_row=s, end_column=10)

        s += 2
        
        min_row = s

        left_bound = 0

        for i, row in df_1.iterrows():

            sheet2['A' + str(s)] = row['bucket']
            sheet2['A' + str(s)].border = thin_border_r
            sheet2['B' + str(s)] = left_bound
            sheet2['B' + str(s)].border = thin_border_r
            sheet2['B' + str(s)].alignment = Alignment(horizontal='right', vertical='center')
            sheet2['C' + str(s)] = round(row['right_bound'] * 100,6)
            sheet2['C' + str(s)].border = thin_border_r
            sheet2['C' + str(s)].alignment = Alignment(horizontal='right', vertical='center')
            sheet2['D' + str(s)] = row['CNT']
            sheet2['D' + str(s)].border = thin_border_r
            sheet2['E' + str(s)] = row['GOODs']
            sheet2['E' + str(s)].border = thin_border_r
            sheet2['F' + str(s)] = row['BADs']
            sheet2['F' + str(s)].border = thin_border_r
            sheet2['G' + str(s)] = row['CNT, %']
            sheet2['G' + str(s)].border = thin_border_r
            sheet2['H' + str(s)] = row['GOODs, %']
            sheet2['H' + str(s)].border = thin_border_r
            sheet2['I' + str(s)] = row['BADs, %']
            sheet2['I' + str(s)].border = thin_border_r
            sheet2['J' + str(s)] = row['Good/OBS, %']
            sheet2['J' + str(s)].border = thin_border_r
            
            sheet2['L' + str(s)] = round(row['GOODs'] * 100 / row['CNT'], 2) if row['CNT'] != 0 else 0
            sheet2['L' + str(s)].border = thin_border_top_bot_l_r

            left_bound = round(row['right_bound'] * 100,6)

            s += 1
        
        max_row = s - 1
        
        ######################
        rule = ColorScaleRule(start_type='min',
                              #start_value=10,
                              start_color='FFF8696B',
                              mid_type='percentile',
                              mid_value=50,
                              mid_color='FFFFEB84', 
                              end_type='max',
                              #end_value=90,
                              end_color='FF63BE7B')
        range_string = 'L' + str(min_row) + ':' + 'L' + str(max_row)
        sheet2.conditional_formatting.add(range_string, rule)
        ######################

        #sheet2['A' + str(s)] = 5
        #sheet2['B' + str(s)] = 'cut-off'
        #sheet2['C' + str(s)] = '=СУММЕСЛИ($A$3:$A$22;">="&$A$23;$C$3:$C$22)'
        #sheet2['D' + str(s)] = '=СУММЕСЛИ($A$3:$A$22;">="&$A$23;$D$3:$D$22)'
        #sheet2['E' + str(s)] = '=C23-D23'
        #sheet2['F' + str(s)] = '=C23/СУММ($C$3:$C$22)'
        #sheet2['G' + str(s)] = '=D23/СУММ($D$3:$D$22)'
        #sheet2['H' + str(s)] = '=E23/СУММ($E$3:$E$22)'
        #sheet2['I' + str(s)] = '=G23/(H23+G23)'
    
        chart1 = BarChart()
        chart1.legend = None
        chart1.height = n_bucket * 0.5 # default is 7.5
        chart1.width = n_bucket * 0.5 * 1.5 # default is 15
        data1 = Reference(sheet2, min_col=8, min_row=min_row, max_row=max_row)
        data2 = Reference(sheet2, min_col=9, min_row=min_row, max_row=max_row)
        chart1.add_data(data1)
        chart1.add_data(data2)
        chart1.y_axis.majorGridlines = None

        c1 = LineChart()
        data3 = Reference(sheet2, min_col=10, min_row=min_row, max_row=max_row)
        c1.add_data(data3)
        c1.y_axis.axId = 200
        s1 = c1.series[0]
        s1.smooth = True

        chart1.y_axis.crosses = "max"
        chart1 += c1

        sheet2.add_chart(chart1, "N" + str(min_row))
        
        s += 4
    
    # Рассчитываем SSI
    obs_sub = round(d_ssi['train'] - d_ssi['test'], 5)
    obs_div = np.where((d_ssi['test'] * d_ssi['train']) != 0.0, d_ssi['train'] / d_ssi['test'], 1.0)
    obs_log = np.log(obs_div)
    
    ssi = round(np.sum(obs_sub * obs_log), 5)
    sheet2['N2'] = 'SSI = ' + str(ssi)
    sheet2['N2'].font = Font(bold=True)


## test_stat

def test_stat(df, df_test, person_id_attr = 'IDCLIENT', target_attr = 'EVENT'):

    # Обрабатываем интервальные переменные

    d_cat = {}

    var_name = list(df[df['int_right_cut'].notna()]['attr_name'].unique())

    for variable in var_name:

        var_test = variable[0:variable.find('$')]

        df_var = df[df['attr_name'] == variable].loc[:,['bin_number', 'int_right_cut']].sort_values(by='bin_number', ascending=True)
        df_var['int_right_cut'] = df_var['int_right_cut'].astype('float64')

        df_1 = df_test.loc[:, [person_id_attr, target_attr, var_test]]
        min_var = np.min(df_1[var_test])
        max_var = np.max(df_1[var_test])
        mean_var = np.round(np.mean(df_1[var_test]), 2)

        condlist = []
        choicelist = []

        for bin_num, right_cut in zip(df_var['bin_number'], df_var['int_right_cut']):
            condlist.append(df_1[var_test] <= right_cut)
            choicelist.append(bin_num)

        df_1[var_test] = np.select(condlist, choicelist)

        d_cat[variable] = 0 if len(df_1[df_1[var_test] == 0]) == 0 else 1

        df_2 = df_1[df_1[var_test] != 0].groupby(by=var_test) \
                                        .aggregate({person_id_attr : 'count', target_attr : 'sum'}) \
                                        .rename(columns={person_id_attr : 'CNT', target_attr : 'quant_1_test'}) \
                                        .reset_index() \
                                        .sort_values(by=var_test, ascending=True)

        row_index = df[df['attr_name'] == variable].index
        df.loc[df.index.isin(row_index) & df['bin_number'].isin(df_2[var_test]), 'quant_0_test'] = df_2['CNT'].values - df_2['quant_1_test'].values
        df.loc[df.index.isin(row_index) & df['bin_number'].isin(df_2[var_test]), 'quant_1_test'] = df_2['quant_1_test'].values


    # Обрабатываем категориальные переменные

    var_cat_name = list(df[df['int_right_cut'].isnull()]['attr_name'].unique())

    for variable in var_cat_name:

        var_test = variable[0:variable.find('$')]

        df_var = df[df['attr_name'] == variable].loc[:,['bin_number', 'cat_attr_list']].sort_values(by='bin_number', ascending=True)

        df_1 = df_test.loc[:, [person_id_attr, target_attr, var_test]]

        condlist = []
        choicelist = []

        for bin_num, cat_attr in zip(df_var['bin_number'], df_var['cat_attr_list']):
            condlist.append(df_1[var_test].isin(cat_attr.split(', ')))
            choicelist.append(bin_num)

        df_1[var_test] = np.select(condlist, choicelist)

        d_cat[variable] = 0 if len(df_1[df_1[var_test] == 0]) == 0 else 1

        df_2 = df_1[df_1[var_test] != 0].groupby(by=var_test) \
                                        .aggregate({person_id_attr : 'count', target_attr : 'sum'}) \
                                        .rename(columns={person_id_attr : 'CNT', target_attr : 'quant_1_test'}) \
                                        .reset_index() \
                                        .sort_values(by=var_test, ascending=True)

        row_index = df[df['attr_name'] == variable].index
        df.loc[df.index.isin(row_index) & df['bin_number'].isin(df_2[var_test]), 'quant_0_test'] = df_2['CNT'].values - df_2['quant_1_test'].values
        df.loc[df.index.isin(row_index) & df['bin_number'].isin(df_2[var_test]), 'quant_1_test'] = df_2['quant_1_test'].values

    # Заполняем пропущенные значения в полях 'quant_0_test', 'quant_1_test'

    df[['quant_0_test', 'quant_1_test']] = df[['quant_0_test', 'quant_1_test']].fillna(0)

    # Приводим поля 'quant_0_test', 'quant_1_test' к типу 'int64'

    df['quant_0_test'] = df['quant_0_test'].astype('int64')
    df['quant_1_test'] = df['quant_1_test'].astype('int64')
    
    for variable in d_cat.keys():
        
        df.loc[(df['attr_name'] == variable) & (df['bin_number'] == 1), 'no category'] = d_cat[variable]
        
    return df


## trend_flg    
        
def trend_flg(df, df_flg, PSI_cutoff = 0.25, PSI_bin_cutoff = False, VOL_cutoff = 0.1, VOL_bin_cutoff = False, bin_v_cutoff = 5.0, break_trd_cutoff = 2):
    
    trend_check = {}

    var_name = list(df['attr_name'].unique())

    for variable in var_name:

        df_4 = df[df['attr_name'] == variable].loc[:, ['bin_number', 'int_left_cut', 'int_right_cut', 'cat_attr_list', 'quant_0', 'quant_1', 'quant_0_test', 'quant_1_test']].sort_values(by='bin_number', ascending=True)
        df_4['cnt_test'] = df_4['quant_1_test'] + df_4['quant_0_test']
        df_4['cnt'] = df_4['quant_1'] + df_4['quant_0']

        all_cnt = df_4['cnt'].sum()
        all_cnt_test = df_4['cnt_test'].sum()

        all_bad = df_4['quant_0'].sum()
        all_good = df_4['quant_1'].sum()

        trend_check[variable] = {'has trend': ['', ''],
                                        'IV': 0,
                                'volatility': [],
                                       'PSI': [],
         'bin-value <'+str(bin_v_cutoff)+'%': [],
                               'no category': df.loc[(df['attr_name'] == variable) 
                                                      & (df['bin_number'] == 1), 'no category'].values[0],
                          'no events in bin': []}

        trend_prev = -1
        trend_prev_test = -1

        for i, row in df_4.iterrows():

            bad_train_rate = row['quant_0'] / all_bad
            good_train_rate = row['quant_1'] / all_good

            E = round(row['cnt'] * 100 / all_cnt, 2)
            I = round(row['cnt_test'] * 100 / all_cnt_test, 2)

            event_per = round(row['quant_1'] * 100 / row['cnt'], 2) if row['cnt'] !=0 else 0

            event_per_test = round(row['quant_1_test'] * 100 / row['cnt_test'], 2) if row['cnt_test'] !=0 else 0
            
            mod_diff = np.absolute((event_per - event_per_test))
            
            weight = np.mean((E, I)) / 100

            bin_psi = round((E - I) * np.log(E / I), 2) if I != 0 else 1.0

            #bin_woe = round(np.log(bad_train_rate / good_train_rate), 6) if good_train_rate != 0 else 0

            bin_iv = round((bad_train_rate - good_train_rate) * np.log(bad_train_rate / good_train_rate), 3) if good_train_rate != 0 else 0

            trend_check[variable]['PSI'].append(bin_psi)
            trend_check[variable]['IV'] += bin_iv
            trend_check[variable]['volatility'].append(mod_diff * weight)
            trend_check[variable]['bin-value <'+str(bin_v_cutoff)+'%'].append(E)
            trend_check[variable]['no events in bin'].append(row['quant_1'])
            trend_check[variable]['has trend'][0] += '1' if event_per > trend_prev else '0'
            trend_check[variable]['has trend'][1] += '1' if event_per_test > trend_prev_test else '0'
            
            trend_prev = event_per
            trend_prev_test = event_per_test

            
    trend_check_2 = {}

    for key in trend_check.keys():
        trend_check_2[key] =    {'has trend': 0,
                            'breaking trend': 0,
                                        'IV': 0,
                                'volatility': 0,
                          'volatility value': 0,
                                       'PSI': 0,
         'bin-value <'+str(bin_v_cutoff)+'%': 0,
                               'no category': trend_check[key]['no category'],
                          'no events in bin': 0}
    
        #trend_check_2[key]['has trend'] = 1 if (trend_check[key]['has trend'][0] == trend_check[key]['has trend'][1]) else 0
        trend_check_2[key]['has trend'] = 1 if sum([i != j for i, j in zip(trend_check[key]['has trend'][0], trend_check[key]['has trend'][1])]) <= break_trd_cutoff else 0
        #trend_check_2[key]['breaking trend'] = sum([i != j for i, j in zip(trend_check[key]['has trend'][0], trend_check[key]['has trend'][1])])
        trend_check_2[key]['IV'] = round(trend_check[key]['IV'], 3)
        
        if VOL_bin_cutoff:
            trend_check_2[key]['volatility'] = 1 if max(trend_check[key]['volatility']) < VOL_cutoff else 0
        else:
            trend_check_2[key]['volatility'] = 1 if sum(trend_check[key]['volatility']) < VOL_cutoff else 0
        
        trend_check_2[key]['volatility value'] = sum(trend_check[key]['volatility'])
        
        if PSI_bin_cutoff:
            trend_check_2[key]['PSI'] = 1 if max(trend_check[key]['PSI']) < PSI_cutoff else 0
        else:
            trend_check_2[key]['PSI'] = 1 if sum(trend_check[key]['PSI']) < PSI_cutoff else 0
        
        trend_check_2[key]['bin-value <'+str(bin_v_cutoff)+'%'] = 1 if min(trend_check[key]['bin-value <'+str(bin_v_cutoff)+'%']) < bin_v_cutoff else 0
        trend_check_2[key]['no events in bin'] = 1 if min(trend_check[key]['no events in bin']) == 0 else 0
    
    for key in trend_check_2.keys():
        df_flg.loc[df_flg['attr_name'] == key, 'has trend'] = trend_check_2[key]['has trend']
        #df_flg.loc[df_flg['attr_name'] == key, 'breaking trends'] = trend_check_2[key]['breaking trend']
        df_flg.loc[df_flg['attr_name'] == key, 'PSI'] = trend_check_2[key]['PSI']
        df_flg.loc[df_flg['attr_name'] == key, 'VOLATILITY'] = trend_check_2[key]['volatility']
        df_flg.loc[df_flg['attr_name'] == key, 'VOLATILITY VALUE'] = trend_check_2[key]['volatility value']
        df_flg.loc[df_flg['attr_name'] == key, 'bin-value <'+str(bin_v_cutoff)+'%'] = trend_check_2[key]['bin-value <'+str(bin_v_cutoff)+'%']
        df_flg.loc[df_flg['attr_name'] == key, 'no category'] = trend_check_2[key]['no category']
        df_flg.loc[df_flg['attr_name'] == key, 'no_events_in_bin'] = trend_check_2[key]['no events in bin']
        
    df_flg.loc[(df_flg['has trend'] == 1) & (df_flg['PSI'] == 1)
                                          & (df_flg['VOLATILITY'] == 1)
                                          & (df_flg['bin-value <'+str(bin_v_cutoff)+'%'] == 0)
                                          & (df_flg['no_events_in_bin'] == 0)
                                          & (df_flg['no category'] == 0), 'trend check'] = 1
    
    df_flg['trend check'] = df_flg['trend check'].fillna(0)
    

## Трендовость + Short_List

def trend(df, df_test, df_coef, wb2, short_lst_, list_size='short', und_r=0, person_id_attr = 'IDCLIENT', target_attr = 'EVENT'):
    
    if list_size == 'short':
        
        sheet_name = 'Trends'
        
        short_lst_name = 'Short-List'
        
        short_lst = short_lst_
        
        
    elif list_size == 'long':
        
        sheet_name = 'Trends (all parameters)'
        
        short_lst_name = 'Short-List (all parameters)'
        
        short_lst_true = [i[0:i.find('$')] for i in short_lst_]
        
        short_lst = []
        for i in short_lst_true:
            for j in list(df['attr_name']):
                if j[0:j.find('$')] == i:
                    short_lst.append(j)
        #short_lst = [i for i in list(df['attr_name']) if i[0:i.find('$')] in short_lst_true]
        #print(short_lst_true)
        #print(short_lst)
    
    # Обрабатываем интервальные переменные

    d_cat = {}

    var_name_all = list(df[df['int_right_cut'].notna()]['attr_name'].unique())
    
    var_name = [i for i in short_lst if i in var_name_all]
    #print(var_name)
    for variable in var_name:
        #print(variable)
        var_test = variable[0:variable.find('$')]

        df_var = df[df['attr_name'] == variable].loc[:,['bin_number', 'int_right_cut']].sort_values(by='bin_number', ascending=True)
        df_var['int_right_cut'] = df_var['int_right_cut'].astype('float64')

        df_1 = df_test.loc[:, [target_attr, var_test]]
        df_1[person_id_attr] = 1
        min_var = np.min(df_1[var_test])
        max_var = np.max(df_1[var_test])
        mean_var = np.round(np.mean(df_1[var_test]), 2)

        condlist = []
        choicelist = []

        for bin_num, right_cut in zip(df_var['bin_number'], df_var['int_right_cut']):
            condlist.append(df_1[var_test] <= right_cut)
            choicelist.append(bin_num)

        df_1[var_test] = np.select(condlist, choicelist)

        d_cat[variable] = {}
        d_cat[variable].setdefault('no category', 0 if len(df_1[df_1[var_test] == 0]) == 0 else 1)
        d_cat[variable].setdefault('CATEGORY', 'INTERVAL')
        #d_cat[variable].setdefault('BIN_CNT', len(df_2))
        d_cat[variable].setdefault('MIN', min_var)
        d_cat[variable].setdefault('MAX', max_var)
        d_cat[variable].setdefault('MEAN', mean_var)

        df_2 = df_1[df_1[var_test] != 0].groupby(by=var_test) \
                                        .aggregate({person_id_attr : 'count', target_attr : 'sum'}) \
                                        .rename(columns={person_id_attr : 'CNT', target_attr : 'quant_1_test'}) \
                                        .reset_index() \
                                        .sort_values(by=var_test, ascending=True)
        
        d_cat[variable].setdefault('BIN_CNT', len(df_2))

        row_index = df[df['attr_name'] == variable].index
        df.loc[df.index.isin(row_index) & df['bin_number'].isin(df_2[var_test]), 'quant_0_test'] = df_2['CNT'].values - df_2['quant_1_test'].values
        df.loc[df.index.isin(row_index) & df['bin_number'].isin(df_2[var_test]), 'quant_1_test'] = df_2['quant_1_test'].values

    
    # Обрабатываем категориальные переменные

    var_cat_name_all = list(df[df['int_right_cut'].isnull()]['attr_name'].unique())
    var_cat_name = [i for i in short_lst if i in var_cat_name_all]

    for variable in var_cat_name:

        var_test = variable[0:variable.find('$')]

        df_var = df[df['attr_name'] == variable].loc[:,['bin_number', 'cat_attr_list']].sort_values(by='bin_number', ascending=True)

        df_1 = df_test.loc[:, [target_attr, var_test]]
        df_1[person_id_attr] = 1

        condlist = []
        choicelist = []

        for bin_num, cat_attr in zip(df_var['bin_number'], df_var['cat_attr_list']):
            condlist.append(df_1[var_test].isin(cat_attr.split(', ')))
            choicelist.append(bin_num)

        df_1[var_test] = np.select(condlist, choicelist)

        d_cat[variable] = {}
        d_cat[variable].setdefault('no category', 0 if len(df_1[df_1[var_test] == 0]) == 0 else 1)
        d_cat[variable].setdefault('CATEGORY', 'CATEGORICAL')
        #d_cat[variable].setdefault('BIN_CNT', len(df_2))
        d_cat[variable].setdefault('MIN', '')
        d_cat[variable].setdefault('MAX', '')
        d_cat[variable].setdefault('MEAN', '')

        df_2 = df_1[df_1[var_test] != 0].groupby(by=var_test) \
                                        .aggregate({person_id_attr : 'count', target_attr : 'sum'}) \
                                        .rename(columns={person_id_attr : 'CNT', target_attr : 'quant_1_test'}) \
                                        .reset_index() \
                                        .sort_values(by=var_test, ascending=True)
        
        d_cat[variable].setdefault('BIN_CNT', len(df_2))

        row_index = df[df['attr_name'] == variable].index
        df.loc[df.index.isin(row_index) & df['bin_number'].isin(df_2[var_test]), 'quant_0_test'] = df_2['CNT'].values - df_2['quant_1_test'].values
        df.loc[df.index.isin(row_index) & df['bin_number'].isin(df_2[var_test]), 'quant_1_test'] = df_2['quant_1_test'].values
    
    # Заполняем пропущенные значения в полях 'quant_0_test', 'quant_1_test'

    df[['quant_0_test', 'quant_1_test']] = df[['quant_0_test', 'quant_1_test']].fillna(0)

    # Приводим поля 'quant_0_test', 'quant_1_test' к типу 'int64'

    df['quant_0_test'] = df['quant_0_test'].astype('int64')
    df['quant_1_test'] = df['quant_1_test'].astype('int64')
    
    
    sheet2 = wb2.create_sheet(sheet_name)
    #sheet2.title = 'Trend'
    sheet2.column_dimensions['A'].width = 50


    trend_check = {}
    
    s = 1

    var_name_all = list(df['attr_name'].unique())
    var_name = [i for i in short_lst if i in var_name_all]

    for variable in var_name:

        df_4 = df[df['attr_name'] == variable].loc[:, ['bin_number', 'int_left_cut', 'int_right_cut', 'cat_attr_list', 'quant_0', 'quant_1', 'quant_0_test', 'quant_1_test']].sort_values(by='bin_number', ascending=True)
        df_4['cnt_test'] = df_4['quant_1_test'] + df_4['quant_0_test']
        df_4['cnt'] = df_4['quant_1'] + df_4['quant_0']

        all_cnt = df_4['cnt'].sum()
        all_cnt_test = df_4['cnt_test'].sum()

        all_bad = df_4['quant_0'].sum()
        all_good = df_4['quant_1'].sum()

        trend_check[variable] = {'has trend': ['', ''],
                                    'IV': 0,
                               'vol_all': 0,
                                'BIN_IV': {},
                                  'GOOD': {},
                                   'BAD': {},
                                   'WOE': {},
                           'volatility' : {},
                                'BOUND' : {},
                                   'PSI': 0,
                         'bin-value <5%': [],
                           'no category': d_cat[variable]['no category'],
                      'no events in bin': [],
                              'CATEGORY': d_cat[variable]['CATEGORY'],
                               'BIN_CNT': d_cat[variable]['BIN_CNT'],
                                   'MIN': d_cat[variable]['MIN'],
                                   'MAX': d_cat[variable]['MAX'],
                                  'MEAN': d_cat[variable]['MEAN']
                            }

        trend_prev = -1
        trend_prev_test = -1
        
        sheet2['A' + str(s)] = variable
        sheet2['A' + str(s)].font = Font(bold=True)
        sheet2['A' + str(s)].border = thin_border_top_bot_r
        sheet2['A' + str(s)].alignment = Alignment(horizontal='center', vertical='center')
        
        sheet2['B' + str(s)].border = thin_border_top_r

        sheet2['C' + str(s)] = 'Bad'
        sheet2['C' + str(s)].font = Font(bold=True)
        sheet2['C' + str(s)].border = thin_border_top_bot
        sheet2['C' + str(s)].alignment = Alignment(horizontal='center', vertical='center')
        sheet2['D' + str(s)] = 'Good'
        sheet2['D' + str(s)].font = Font(bold=True)
        sheet2['D' + str(s)].border = thin_border_top_bot
        sheet2['D' + str(s)].alignment = Alignment(horizontal='center', vertical='center')
        sheet2['E' + str(s)].border = thin_border_top_bot
        sheet2['F' + str(s)].border = thin_border_top_bot_r
        sheet2['G' + str(s)] = 'Bad'
        sheet2['G' + str(s)].font = Font(bold=True)
        sheet2['G' + str(s)].border = thin_border_top_bot
        sheet2['G' + str(s)].alignment = Alignment(horizontal='center', vertical='center')
        sheet2['H' + str(s)] = 'Good'
        sheet2['H' + str(s)].font = Font(bold=True)
        sheet2['H' + str(s)].border = thin_border_top_bot
        sheet2['H' + str(s)].alignment = Alignment(horizontal='center', vertical='center')
        sheet2['I' + str(s)].border = thin_border_top_bot
        sheet2['J' + str(s)].border = thin_border_top_bot_r
        sheet2['K' + str(s)] = 'PSI'
        sheet2['K' + str(s)].font = Font(bold=True)
        sheet2['K' + str(s)].border = thin_border_top_bot_r
        sheet2['K' + str(s)].alignment = Alignment(horizontal='center', vertical='center')
        sheet2['L' + str(s)].border = thin_border_top_bot_r
        sheet2['M' + str(s)] = 'd_bad_p'
        #sheet2['M' + str(s)].font = Font(bold=True)
        sheet2['M' + str(s)].border = thin_border_top_bot
        sheet2['M' + str(s)].alignment = Alignment(horizontal='center', vertical='center')
        sheet2['N' + str(s)] = 'd_good_p'
        #sheet2['N' + str(s)].font = Font(bold=True)
        sheet2['N' + str(s)].border = thin_border_top_bot
        sheet2['N' + str(s)].alignment = Alignment(horizontal='center', vertical='center')
        sheet2['O' + str(s)] = 'WOE'
        sheet2['O' + str(s)].font = Font(bold=True)
        sheet2['O' + str(s)].border = thin_border_top_bot
        sheet2['O' + str(s)].alignment = Alignment(horizontal='center', vertical='center')
        sheet2['P' + str(s)] = 'IV'
        sheet2['P' + str(s)].font = Font(bold=True)
        sheet2['P' + str(s)].border = thin_border_top_bot_r
        sheet2['P' + str(s)].alignment = Alignment(horizontal='center', vertical='center')
        sheet2['Q' + str(s)] = 'SCOR'
        sheet2['Q' + str(s)].font = Font(bold=True)
        sheet2['Q' + str(s)].border = thin_border_top_bot_r
        sheet2['Q' + str(s)].alignment = Alignment(horizontal='center', vertical='center')
        sheet2['R' + str(s)] = 'NUM_FROM'
        sheet2['R' + str(s)].font = Font(bold=True)
        sheet2['R' + str(s)].border = thin_border_top_bot
        sheet2['R' + str(s)].alignment = Alignment(horizontal='center', vertical='center')
        sheet2['S' + str(s)] = 'NUM_TO'
        sheet2['S' + str(s)].font = Font(bold=True)
        sheet2['S' + str(s)].border = thin_border_top_bot_r
        sheet2['S' + str(s)].alignment = Alignment(horizontal='center', vertical='center')
        sheet2['T' + str(s)] = 'RANGE'
        sheet2['T' + str(s)].font = Font(bold=True)
        sheet2['T' + str(s)].border = thin_border_top_bot_r
        sheet2['T' + str(s)].alignment = Alignment(horizontal='center', vertical='center')
        
        len_d = df_4['bin_number'].max()
        
        corr_coef = df_coef.loc[df_coef['ATTR_TRUE_NAME'] == variable[0:variable.find('$')], 'COEF'].values[0]
        
        sheet2.merge_cells(start_row=s, start_column=1, end_row=s+len_d, end_column=1)
        
        for i, row in df_4.iterrows():
           
            s += 1

            bad_train_rate = row['quant_0'] / all_bad
            good_train_rate = row['quant_1'] / all_good

            E = round(row['cnt'] * 100 / all_cnt, 2)
            I = round(row['cnt_test'] * 100 / all_cnt_test, 2)

            event_per = round(row['quant_1'] * 100 / row['cnt'], 2) if row['cnt'] !=0 else 0

            event_per_test = round(row['quant_1_test'] * 100 / row['cnt_test'], 2) if row['cnt_test'] !=0 else 0
            
            mod_diff = np.absolute((event_per - event_per_test))
            
            weight = np.mean((E, I)) / 100

            bin_psi = round((E - I) * np.log(E / I), 2) if I != 0 else 1.0

            bin_woe = round(np.log(bad_train_rate / good_train_rate), 6) if good_train_rate != 0 else 0

            bin_iv = round((bad_train_rate - good_train_rate) * np.log(bad_train_rate / good_train_rate), 3) if good_train_rate != 0 else 0
            
            sheet2['B' + str(s)] = row['bin_number']
            sheet2['C' + str(s)] = row['quant_0']
            sheet2['D' + str(s)] = row['quant_1']
            sheet2['E' + str(s)] = event_per
            sheet2['F' + str(s)] = E

            sheet2['G' + str(s)] = row['quant_0_test']
            sheet2['H' + str(s)] = row['quant_1_test']
            sheet2['I' + str(s)] = event_per_test
            sheet2['J' + str(s)] = I

            sheet2['K' + str(s)] = bin_psi

            sheet2['M' + str(s)] = round(bad_train_rate, 6)
            sheet2['N' + str(s)] = round(good_train_rate, 6)
            sheet2['O' + str(s)] = bin_woe
            sheet2['P' + str(s)] = bin_iv
            sheet2['Q' + str(s)] = round(bin_woe * corr_coef, 6)
            sheet2['R' + str(s)] = row['int_left_cut'] if  (str(row['cat_attr_list']).lower() == 'nan' or str(row['cat_attr_list']).lower() == '') else ''
            sheet2['S' + str(s)] = row['int_right_cut'] if (str(row['cat_attr_list']).lower() == 'nan' or str(row['cat_attr_list']).lower() == '') else ''
            sheet2['T' + str(s)] = row['cat_attr_list'] if (str(row['cat_attr_list']).lower() != 'nan' or str(row['cat_attr_list']).lower() != '') else ''
            
            sheet2['A' + str(s)].border = thin_border_r
            sheet2['B' + str(s)].border = thin_border_r
            sheet2['F' + str(s)].border = thin_border_r
            sheet2['J' + str(s)].border = thin_border_r
            sheet2['K' + str(s)].border = thin_border_r
            sheet2['L' + str(s)].border = thin_border_r
            sheet2['P' + str(s)].border = thin_border_r
            sheet2['Q' + str(s)].border = thin_border_r
            sheet2['S' + str(s)].border = thin_border_r
            sheet2['T' + str(s)].border = thin_border_r
            
            trend_check[variable]['PSI'] += bin_psi
            trend_check[variable]['IV'] += bin_iv
            trend_check[variable]['vol_all'] += mod_diff * weight
            trend_check[variable]['BIN_IV'].setdefault(row['bin_number'], bin_iv)
            trend_check[variable]['GOOD'].setdefault(row['bin_number'], row['quant_1'])
            trend_check[variable]['BAD'].setdefault(row['bin_number'], row['quant_0'])
            trend_check[variable]['WOE'].setdefault(row['bin_number'], bin_woe)
            trend_check[variable]['volatility'].setdefault(row['bin_number'], mod_diff * weight)
            trend_check[variable]['BOUND'].setdefault(row['bin_number'], {'NUM_FROM' : row['int_left_cut'], 'NUM_TO' : row['int_right_cut']}) if (str(row['cat_attr_list']).lower() == 'nan' or str(row['cat_attr_list']).lower() == '') else trend_check[variable]['BOUND'].setdefault(row['bin_number'], {'RANGE': row['cat_attr_list']})
            trend_check[variable]['bin-value <5%'].append(E)
            trend_check[variable]['no events in bin'].append(row['quant_1'])
            trend_check[variable]['has trend'][0] += '1' if event_per > trend_prev else '0'
            trend_check[variable]['has trend'][1] += '1' if event_per_test > trend_prev_test else '0'
            
            trend_prev = event_per
            trend_prev_test = event_per_test
            
        sheet2['A' + str(s)].border = thin_border_bot_r
        sheet2['B' + str(s)].border = thin_border_bot_r
        sheet2['C' + str(s)].border = thin_border_bot
        sheet2['D' + str(s)].border = thin_border_bot
        sheet2['K' + str(s)].border = thin_border_bot
        sheet2['E' + str(s)].border = thin_border_bot
        sheet2['F' + str(s)].border = thin_border_bot_r
        sheet2['G' + str(s)].border = thin_border_bot
        sheet2['H' + str(s)].border = thin_border_bot
        sheet2['I' + str(s)].border = thin_border_bot
        sheet2['J' + str(s)].border = thin_border_bot_r
        sheet2['K' + str(s)].border = thin_border_bot_r
        sheet2['L' + str(s)].border = thin_border_bot_r
        sheet2['M' + str(s)].border = thin_border_bot
        sheet2['N' + str(s)].border = thin_border_bot
        sheet2['O' + str(s)].border = thin_border_bot
        sheet2['P' + str(s)].border = thin_border_bot_r
        sheet2['Q' + str(s)].border = thin_border_bot_r
        sheet2['R' + str(s)].border = thin_border_bot
        sheet2['S' + str(s)].border = thin_border_bot_r
        sheet2['T' + str(s)].border = thin_border_bot_r
        
        chart1 = BarChart()
        chart1.legend = None
        data1 = Reference(sheet2, min_col=6, min_row=s-len_d+1, max_row=s)
        data2 = Reference(sheet2, min_col=10, min_row=s-len_d+1, max_row=s)
        chart1.add_data(data1)
        chart1.add_data(data2)
        chart1.y_axis.majorGridlines = None

        c1 = LineChart()
        data3 = Reference(sheet2, min_col=5, min_row=s-len_d+1, max_row=s)
        data4 = Reference(sheet2, min_col=9, min_row=s-len_d+1, max_row=s)
        c1.add_data(data3)
        c1.add_data(data4)
        c1.y_axis.axId = 200

        chart1.y_axis.crosses = "max"
        chart1 += c1

        sheet2.add_chart(chart1, "A" + str(s+1))

        s += 17
    
    trend_check_2 = {}
    for key in trend_check.keys():
        trend_check_2[key] =    {'has trend': 0,
                                        'IV': 0,
                                   'vol_all': 0,
                                    'BIN_IV': trend_check[key]['BIN_IV'],
                                      'GOOD': trend_check[key]['GOOD'],
                                       'BAD': trend_check[key]['BAD'],
                                       'WOE': trend_check[key]['WOE'],
                               'volatility' : trend_check[key]['volatility'],
                                     'BOUND': trend_check[key]['BOUND'],
                                       'PSI': 0,
                             'bin-value <5%': 0,
                               'no category': trend_check[key]['no category'],
                          'no events in bin': 0,
                                  'CATEGORY': trend_check[key]['CATEGORY'],
                                   'BIN_CNT': trend_check[key]['BIN_CNT'],
                                       'MIN': trend_check[key]['MIN'],
                                       'MAX': trend_check[key]['MAX'],
                                      'MEAN': trend_check[key]['MEAN']
                                }
    
        trend_check_2[key]['has trend'] = 1 if trend_check[key]['has trend'][0] == trend_check[key]['has trend'][1] else 0
        trend_check_2[key]['IV'] = round(trend_check[key]['IV'], 3)
        trend_check_2[key]['vol_all'] = trend_check[key]['vol_all']
        trend_check_2[key]['PSI'] = trend_check[key]['PSI']
        trend_check_2[key]['bin-value <5%'] = 1 if min(trend_check[key]['bin-value <5%']) < 5.0 else 0
        trend_check_2[key]['no events in bin'] = 1 if min(trend_check[key]['no events in bin']) == 0 else 0
    
    # Составляем скоркарту
    
    sc_strings = []
    
    intercept = df_coef.loc[df_coef['ATTR_TRUE_NAME'] == 'INTERCEPT', 'COEF'].values[0]
    string_2 = 'cast(1/(1 + exp(-(' + str(round(intercept, 6))
    string_3 = ''
    string_4 = ''
    
    if und_r > 0:
        
        string_4 += 'round((t.score * (1/' + str(und_r) + ')) / (t.score * (1/' + str(und_r) + ') - t.score + 1), 6) as score'
    
    for key in trend_check_2.keys():
        
        v =  key[0:key.find('$')]
       
        string_2 += '\n           + t.' + v
        
        corr_coef = df_coef.loc[df_coef['ATTR_TRUE_NAME'] == v, 'COEF'].values[0]
        
        string = 'case'
        kk = trend_check_2[key]['BIN_CNT']
        
        if trend_check_2[key]['CATEGORY'] == 'INTERVAL':
            
            string_3 += 'nvl(t.' + v + ', -1) as ' + v + ',' + '\n'
        
            for k in range(1, kk):

                string += '\n    when ' + 't.' + v + ' <= ' + str(trend_check_2[key]['BOUND'][k]['NUM_TO']) + ' then ' + str(round(trend_check_2[key]['WOE'][k] * corr_coef, 6))

            string += '\n    else ' + str(round(trend_check_2[key]['WOE'][kk] * corr_coef, 6))
            string += '\nend as ' + v + ','

            sc_strings.append(string)
        
        else:
            
            string_3 += 'nvl(t.' + v + ', "EMPTY") as ' + v + ',' + '\n'
            
            for k in range(1, kk + 1):
                string += '\n    when ' + 't.' + v + ' in (' + str(trend_check_2[key]['BOUND'][k]['RANGE']) + ') then ' + str(round(trend_check_2[key]['WOE'][k] * corr_coef, 6))
            
            string += '\n    else ' + str(0)
            string += '\nend as ' + v + ','
            
            sc_strings.append(string)
       
    scor_card = '\n\n'.join(sc_strings)
    
    string_2 += '))) as decimal(18, 2)) as score'
    
     # Составляем Short-List
    sheet3 = wb2.create_sheet(short_lst_name) 
    sheet3.column_dimensions['B'].width = 50
    sheet3.column_dimensions['C'].width = 15
    sheet3.column_dimensions['D'].width = 15
    sheet3.column_dimensions['I'].width = 15
    sheet3.column_dimensions['J'].width = 15
    sheet3.column_dimensions['K'].width = 15
    
    sheet3['R2'] = string_3
    
    sheet3['R4'] = scor_card
    
    sheet3['R6'] = string_2
    
    sheet3['R8'] = string_4
    
    #df_fields = load_data_from_oracle_2('crm_score.lib_data')[['DATA_FIELD', 'DESCRIBE']]
    #df_fields['DESCRIBE'] = df_fields['DESCRIBE'].astype('str')
    
    s = 1
    for var_n, key in enumerate(trend_check_2.keys()):

        sheet3['A' + str(s)] = '#'
        sheet3['A' + str(s)].fill = PatternFill(start_color='00969696', fill_type = "solid")
        sheet3['A' + str(s)].font = Font(bold=True)
        sheet3['A' + str(s)].border = thin_border_top_bot_r
        sheet3['A' + str(s)].alignment = Alignment(horizontal='center', vertical='center')
        sheet3['A' + str(s+1)] = var_n + 1
        sheet3['A' + str(s+1)].alignment = Alignment(horizontal='center', vertical='center')

        sheet3['B' + str(s)] = 'NAME'
        sheet3['B' + str(s)].fill = PatternFill(start_color='00969696', fill_type = "solid")
        sheet3['B' + str(s)].font = Font(bold=True)
        sheet3['B' + str(s)].border = thin_border_top_bot_r
        sheet3['B' + str(s)].alignment = Alignment(horizontal='center', vertical='center')
        #sheet3['B' + str(s+1)] = df_fields.loc[df_fields['DATA_FIELD'] == key[0:key.find('$')], 'DESCRIBE'].values[0]
        sheet3['B' + str(s+1)].alignment = Alignment(horizontal='left', vertical='center')
        sheet3['B' + str(s+3)] = key
        sheet3['B' + str(s+3)].alignment = Alignment(horizontal='center', vertical='center')

        sheet3['C' + str(s)] = 'CATEGORY'
        sheet3['C' + str(s)].fill = PatternFill(start_color='00969696', fill_type = "solid")
        sheet3['C' + str(s)].font = Font(bold=True)
        sheet3['C' + str(s)].border = thin_border_top_bot_r
        sheet3['C' + str(s)].alignment = Alignment(horizontal='center', vertical='center')
        sheet3['C' + str(s+1)] = trend_check_2[key]['CATEGORY']
        sheet3['C' + str(s+1)].border = thin_border_r
        sheet3['C' + str(s+1)].alignment = Alignment(horizontal='center', vertical='center')
        sheet3['C' + str(s+2)] = '#'
        sheet3['C' + str(s+2)].fill = PatternFill(start_color='00C0C0C0', fill_type = "solid")
        sheet3['C' + str(s+2)].border = thin_border_top_bot_r
        sheet3['C' + str(s+2)].alignment = Alignment(horizontal='center', vertical='center')
        
        sheet3['D' + str(s)] = 'VOLATILITY'
        sheet3['D' + str(s)].fill = PatternFill(start_color='00969696', fill_type = "solid")
        sheet3['D' + str(s)].font = Font(bold=True)
        sheet3['D' + str(s)].border = thin_border_top_bot_r
        sheet3['D' + str(s)].alignment = Alignment(horizontal='center', vertical='center')
        sheet3['D' + str(s+1)] = trend_check_2[key]['vol_all']
        sheet3['D' + str(s+1)].border = thin_border_r
        sheet3['D' + str(s+1)].alignment = Alignment(horizontal='center', vertical='center')
        sheet3['D' + str(s+2)] = 'BIN_VOL'
        sheet3['D' + str(s+2)].fill = PatternFill(start_color='00C0C0C0', fill_type = "solid")
        sheet3['D' + str(s+2)].border = thin_border_top_bot_r
        sheet3['D' + str(s+2)].alignment = Alignment(horizontal='center', vertical='center')
        
        sheet3['E' + str(s)] = 'SOURCE'
        sheet3['E' + str(s)].fill = PatternFill(start_color='00969696', fill_type = "solid")
        sheet3['E' + str(s)].font = Font(bold=True)
        sheet3['E' + str(s)].border = thin_border_top_bot_r
        sheet3['E' + str(s)].alignment = Alignment(horizontal='center', vertical='center')
        sheet3['E' + str(s+1)] = key[0:key.find('$')]
        sheet3['E' + str(s+1)].alignment = Alignment(horizontal='center', vertical='center')
        sheet3['E' + str(s+2)] = 'WOE'
        sheet3['E' + str(s+2)].fill = PatternFill(start_color='00C0C0C0', fill_type = "solid")
        sheet3['E' + str(s+2)].border = thin_border_top_bot_r
        sheet3['E' + str(s+2)].alignment = Alignment(horizontal='center', vertical='center')

        sheet3['F' + str(s)] = 'IV'
        sheet3['F' + str(s)].fill = PatternFill(start_color='00969696', fill_type = "solid")
        sheet3['F' + str(s)].font = Font(bold=True)
        sheet3['F' + str(s)].border = thin_border_top_bot_r
        sheet3['F' + str(s)].alignment = Alignment(horizontal='center', vertical='center')
        sheet3['F' + str(s+1)] = trend_check_2[key]['IV']
        sheet3['F' + str(s+1)].border = thin_border_r
        sheet3['F' + str(s+1)].alignment = Alignment(horizontal='center', vertical='center')
        sheet3['F' + str(s+2)] = 'BIN_IV'
        sheet3['F' + str(s+2)].fill = PatternFill(start_color='00C0C0C0', fill_type = "solid")
        sheet3['F' + str(s+2)].border = thin_border_top_bot_r
        sheet3['F' + str(s+2)].alignment = Alignment(horizontal='center', vertical='center')

        sheet3['G' + str(s)] = 'PSI'
        sheet3['G' + str(s)].fill = PatternFill(start_color='00969696', fill_type = "solid")
        sheet3['G' + str(s)].font = Font(bold=True)
        sheet3['G' + str(s)].border = thin_border_top_bot_r
        sheet3['G' + str(s)].alignment = Alignment(horizontal='center', vertical='center')
        sheet3['G' + str(s+1)] = trend_check_2[key]['PSI']
        sheet3['G' + str(s+1)].border = thin_border_r
        sheet3['G' + str(s+1)].alignment = Alignment(horizontal='center', vertical='center')
        sheet3['G' + str(s+2)] = 'GOOD'
        sheet3['G' + str(s+2)].fill = PatternFill(start_color='00C0C0C0', fill_type = "solid")
        sheet3['G' + str(s+2)].border = thin_border_top_bot_r
        sheet3['G' + str(s+2)].alignment = Alignment(horizontal='center', vertical='center')

        sheet3['H' + str(s)] = 'TREND'
        sheet3['H' + str(s)].fill = PatternFill(start_color='00969696', fill_type = "solid")
        sheet3['H' + str(s)].font = Font(bold=True)
        sheet3['H' + str(s)].border = thin_border_top_bot_r
        sheet3['H' + str(s)].alignment = Alignment(horizontal='center', vertical='center')
        sheet3['H' + str(s+1)] = trend_check_2[key]['has trend']
        sheet3['H' + str(s+1)].border = thin_border_r
        sheet3['H' + str(s+1)].alignment = Alignment(horizontal='center', vertical='center')
        sheet3['H' + str(s+2)] = 'BAD'
        sheet3['H' + str(s+2)].fill = PatternFill(start_color='00C0C0C0', fill_type = "solid")
        sheet3['H' + str(s+2)].border = thin_border_top_bot_r
        sheet3['H' + str(s+2)].alignment = Alignment(horizontal='center', vertical='center')

        sheet3['I' + str(s)] = 'BIN_CNT'
        sheet3['I' + str(s)].fill = PatternFill(start_color='00969696', fill_type = "solid")
        sheet3['I' + str(s)].font = Font(bold=True)
        sheet3['I' + str(s)].border = thin_border_top_bot_r
        sheet3['I' + str(s)].alignment = Alignment(horizontal='center', vertical='center')
        sheet3['I' + str(s+1)] = trend_check_2[key]['BIN_CNT']
        sheet3['I' + str(s+1)].border = thin_border_r
        sheet3['I' + str(s+1)].alignment = Alignment(horizontal='center', vertical='center')
        sheet3['I' + str(s+2)] = 'SCOR'
        sheet3['I' + str(s+2)].fill = PatternFill(start_color='00C0C0C0', fill_type = "solid")
        sheet3['I' + str(s+2)].border = thin_border_top_bot_r
        sheet3['I' + str(s+2)].alignment = Alignment(horizontal='center', vertical='center')

        sheet3['J' + str(s)] = 'MIN'
        sheet3['J' + str(s)].fill = PatternFill(start_color='00969696', fill_type = "solid")
        sheet3['J' + str(s)].font = Font(bold=True)
        sheet3['J' + str(s)].border = thin_border_top_bot_r
        sheet3['J' + str(s)].alignment = Alignment(horizontal='center', vertical='center')
        sheet3['J' + str(s+1)] = trend_check_2[key]['MIN']
        sheet3['J' + str(s+1)].border = thin_border_r
        sheet3['J' + str(s+1)].alignment = Alignment(horizontal='center', vertical='center')
        sheet3['J' + str(s+2)] = 'NUM_FROM'
        sheet3['J' + str(s+2)].fill = PatternFill(start_color='00C0C0C0', fill_type = "solid")
        sheet3['J' + str(s+2)].border = thin_border_top_bot_r
        sheet3['J' + str(s+2)].alignment = Alignment(horizontal='center', vertical='center')

        sheet3['K' + str(s)] = 'MAX'
        sheet3['K' + str(s)].fill = PatternFill(start_color='00969696', fill_type = "solid")
        sheet3['K' + str(s)].font = Font(bold=True)
        sheet3['K' + str(s)].border = thin_border_top_bot_r
        sheet3['K' + str(s)].alignment = Alignment(horizontal='center', vertical='center')
        sheet3['K' + str(s+1)] = trend_check_2[key]['MAX']
        sheet3['K' + str(s+1)].border = thin_border_r
        sheet3['K' + str(s+1)].alignment = Alignment(horizontal='center', vertical='center')
        sheet3['K' + str(s+2)] = 'NUM_TO'
        sheet3['K' + str(s+2)].fill = PatternFill(start_color='00C0C0C0', fill_type = "solid")
        sheet3['K' + str(s+2)].border = thin_border_top_bot_r
        sheet3['K' + str(s+2)].alignment = Alignment(horizontal='center', vertical='center')

        sheet3['L' + str(s)] = 'MEAN'
        sheet3['L' + str(s)].fill = PatternFill(start_color='00969696', fill_type = "solid")
        sheet3['L' + str(s)].font = Font(bold=True)
        sheet3['L' + str(s)].border = thin_border_top_bot_r
        sheet3['L' + str(s)].alignment = Alignment(horizontal='center', vertical='center')
        sheet3['L' + str(s+1)] = trend_check_2[key]['MEAN']
        sheet3['L' + str(s+1)].border = thin_border_r
        sheet3['L' + str(s+1)].alignment = Alignment(horizontal='center', vertical='center')
        sheet3['L' + str(s+2)] = 'RANGE'
        sheet3['L' + str(s+2)].fill = PatternFill(start_color='00C0C0C0', fill_type = "solid")
        sheet3['L' + str(s+2)].border = thin_border_top_bot_r
        sheet3['L' + str(s+2)].alignment = Alignment(horizontal='center', vertical='center')

        len_d = trend_check_2[key]['BIN_CNT']
        sheet3.merge_cells(start_row=s+1, start_column=1, end_row=s+2+len_d, end_column=1)
        sheet3.merge_cells(start_row=s+1, start_column=2, end_row=s+2, end_column=2)
        sheet3.merge_cells(start_row=s+3, start_column=2, end_row=s+2+len_d, end_column=2)
        
        sheet3['A' + str(s+1)].border = thin_border_r
        sheet3['A' + str(s+2)].border = thin_border_r
        sheet3['B' + str(s+1)].border = thin_border_r
        sheet3['B' + str(s+2)].border = thin_border_r
        
        corr_coef = df_coef.loc[df_coef['ATTR_TRUE_NAME'] == key[0:key.find('$')], 'COEF'].values[0]
        
        s += 3
        for j in range(1, trend_check_2[key]['BIN_CNT'] + 1):
            
            sheet3['A' + str(s)].border = thin_border_r
            sheet3['B' + str(s)].border = thin_border_r
            sheet3['C' + str(s)] = j
            sheet3['C' + str(s)].border = thin_border_r
            sheet3['C' + str(s)].alignment = Alignment(horizontal='center', vertical='center')
            sheet3['D' + str(s)] = trend_check_2[key]['volatility'][j]
            sheet3['D' + str(s)].border = thin_border_r
            sheet3['E' + str(s)] = trend_check_2[key]['WOE'][j]
            sheet3['E' + str(s)].border = thin_border_r
            sheet3['F' + str(s)] = trend_check_2[key]['BIN_IV'][j]
            sheet3['F' + str(s)].border = thin_border_r
            sheet3['G' + str(s)] = trend_check_2[key]['GOOD'][j]
            sheet3['G' + str(s)].border = thin_border_r
            sheet3['H' + str(s)] = trend_check_2[key]['BAD'][j]
            sheet3['H' + str(s)].border = thin_border_r
            sheet3['I' + str(s)] = round(trend_check_2[key]['WOE'][j] * corr_coef, 6)
            sheet3['I' + str(s)].border = thin_border_r
            sheet3['J' + str(s)] = trend_check_2[key]['BOUND'][j]['NUM_FROM'] if trend_check_2[key]['CATEGORY'] == 'INTERVAL' else ''
            sheet3['J' + str(s)].border = thin_border_r
            sheet3['K' + str(s)] = trend_check_2[key]['BOUND'][j]['NUM_TO'] if trend_check_2[key]['CATEGORY'] == 'INTERVAL' else ''
            sheet3['K' + str(s)].border = thin_border_r
            sheet3['L' + str(s)] = trend_check_2[key]['BOUND'][j]['RANGE'] if trend_check_2[key]['CATEGORY'] == 'CATEGORICAL' else ''
            sheet3['L' + str(s)].border = thin_border_r
            
            s += 1
        
        sheet3['A' + str(s)].border = thin_border_top
        sheet3['B' + str(s)].border = thin_border_top
        sheet3['C' + str(s)].border = thin_border_top
        sheet3['D' + str(s)].border = thin_border_top
        sheet3['E' + str(s)].border = thin_border_top
        sheet3['F' + str(s)].border = thin_border_top
        sheet3['G' + str(s)].border = thin_border_top
        sheet3['H' + str(s)].border = thin_border_top
        sheet3['I' + str(s)].border = thin_border_top
        sheet3['J' + str(s)].border = thin_border_top
        sheet3['K' + str(s)].border = thin_border_top
        sheet3['L' + str(s)].border = thin_border_top
            
        s += 1
  
